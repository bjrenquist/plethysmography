import csv
import openpyxl
# noinspection PyUnresolvedReferences
import lxml
import numpy as np
from scipy import interpolate, stats
from scipy.signal import butter, sosfiltfilt
from scipy.ndimage import gaussian_filter1d
from matplotlib import pyplot as plt


class Respiration:

    def __init__(self):
        # Data Signals: each element represents one data point, collected at the specified sampling_rate
        self.data_idx = np.array([])
        self.time = np.array([])
        self.comments = np.array([])
        self.voltage = np.array([])
        self.pressure = np.array([])
        self.lowpass_pressure = np.array([])
        self.volume = np.array([])
        self.smoothed_volume = np.array([])
        self.global_corrected_volume = np.array([])
        self.local_corrected_volume = np.array([])
        self.flow = np.array([])
        self.smoothed_flow = np.array([])

        # Data Landmarks: each element represents the index at which the landmark occurs within the data signals
        self.consensus_peaks = np.ma.array([])
        self.consensus_troughs = np.ma.array([])
        self.consensus_all_peaks = []
        self.consensus_all_troughs = []
        self.min_inhale_volume_idx = np.array([])
        self.max_inhale_exhale_volume_idx = np.array([])
        self.min_exhale_volume_idx = np.array([])
        self.final_peak = np.array([])
        self.initial_trough = np.array([])
        # self.consensus_first_peaks = np.ma.array([])
        # self.consensus_first_troughs = np.ma.array([])
        # self.consensus_last_peaks = np.ma.array([])
        # self.consensus_last_troughs = np.ma.array([])
        # self.initial_trough_first = int()
        # self.initial_trough_last = int()
        # self.initial_trough_min = int()
        # self.final_peak_first = int()
        # self.final_peak_last = int()
        # self.final_peak_max = int()
        self.inhale_onsets = np.ma.array([])
        self.inhale_offsets = np.ma.array([])
        self.inhale_pause_onsets = np.ma.array([])
        self.inhale_pause_offsets = np.ma.array([])
        self.exhale_onsets = np.ma.array([])
        self.exhale_offsets = np.ma.array([])
        self.exhale_pause_onsets = np.ma.array([])
        self.exhale_pause_offsets = np.ma.array([])
        self.inhale_flow_peaks = np.ma.array([])
        self.exhale_flow_peaks = np.ma.array([])
        self.inhale_volume_peaks = np.ma.array([])
        self.exhale_volume_peaks = np.ma.array([])

        self.start_idx = np.ma.array([])
        self.stop_idx = np.ma.array([])
        self.i50v = np.ma.array([])
        self.i64v = np.ma.array([])
        self.e50v = np.ma.array([])
        self.e64v = np.ma.array([])

        # Stats; use masked arrays to later exclude breaths from analysis
        self.inhale_volumes = np.ma.array([])
        self.inhale_pause_volumes = np.ma.array([])
        self.exhale_volumes = np.ma.array([])
        self.exhale_pause_volumes = np.ma.array([])
        self.inhale_pause_plus_exhale_volumes = np.ma.array([])
        self.inhale_peak_flows = np.ma.array([])
        self.exhale_peak_flows = np.ma.array([])
        self.if50 = np.ma.array([])
        self.ef50 = np.ma.array([])
        self.inhale_times = np.ma.array([])
        self.inhale_pause_times = np.ma.array([])
        self.exhale_times = np.ma.array([])
        self.exhale_pause_times = np.ma.array([])
        self.inhale_pause_plus_exhale_times = np.ma.array([])
        self.breath_times = np.ma.array([])
        self.bpm = np.ma.array([])
        self.tri_50 = np.ma.array([])
        self.tri_64 = np.ma.array([])
        self.tre_50 = np.ma.array([])
        self.tre_64 = np.ma.array([])
        self.penh = np.ma.array([])
        self.saturated = np.array([])
        self.movement = np.array([])
        self.outlier = np.array([])
        self.outlier_piv = np.array([])
        self.outlier_pev = np.array([])
        self.outlier_pif = np.array([])
        self.outlier_pef = np.array([])
        self.outlier_ti = np.array([])
        self.outlier_te = np.array([])
        self.outlier_tb = np.array([])
        self.proximity = np.array([])
        self.breaths_to_exclude = np.array([])

        # Stat Avgs
        self.avg_inhale_volume = np.array([])
        self.avg_inhale_pause_volume = np.array([])
        self.avg_exhale_volume = np.array([])
        self.avg_exhale_pause_volume = np.array([])
        self.avg_inhale_pause_plus_exhale_volume = np.array([])
        self.avg_inhale_peak_flow = np.array([])
        self.avg_exhale_peak_flow = np.array([])
        self.avg_if50 = np.array([])
        self.avg_ef50 = np.array([])
        self.avg_inhale_time = np.array([])
        self.avg_inhale_pause_time = np.array([])
        self.avg_exhale_time = np.array([])
        self.avg_exhale_pause_time = np.array([])
        self.avg_inhale_pause_plus_exhale_time = np.array([])
        self.avg_breath_time = np.array([])
        self.avg_bpm = np.array([])
        self.avg_tri_50 = np.array([])
        self.avg_tri_64 = np.array([])
        self.avg_tre_50 = np.array([])
        self.avg_tre_64 = np.array([])
        self.avg_penh = np.array([])
        self.stdev_inhale_volume = np.array([])
        self.stdev_inhale_pause_volume = np.array([])
        self.stdev_exhale_volume = np.array([])
        self.stdev_exhale_pause_volume = np.array([])
        self.stdev_inhale_pause_plus_exhale_volume = np.array([])
        self.stdev_inhale_peak_flow = np.array([])
        self.stdev_exhale_peak_flow = np.array([])
        self.stdev_if50 = np.array([])
        self.stdev_ef50 = np.array([])
        self.stdev_inhale_time = np.array([])
        self.stdev_inhale_pause_time = np.array([])
        self.stdev_exhale_time = np.array([])
        self.stdev_exhale_pause_time = np.array([])
        self.stdev_inhale_pause_plus_exhale_time = np.array([])
        self.stdev_breath_time = np.array([])
        self.stdev_bpm = np.array([])
        self.stdev_tri_50 = np.array([])
        self.stdev_tri_64 = np.array([])
        self.stdev_tre_50 = np.array([])
        self.stdev_tre_64 = np.array([])
        self.stdev_penh = np.array([])
        self.window_start_time = np.array([])
        self.window_stop_time = np.array([])
        self.window_rel_center_time = np.array([])
        self.window_num_breaths = np.array([])
        self.window_num_excluded = np.array([])
        self.window_num_saturated = np.array([])
        self.window_num_movement = np.array([])
        self.window_num_outlier = np.array([])

        # Data Collection Parameters
        self.data_zero = float()
        self.len_data = int()
        self.num_breaths = int()
        self.num_windows = int()
        self.channel_count = int()
        self.sample_count = int()
        self.sampling_rate = float()
        self.start_time = float()
        self.comment_idxs = np.array([])
        self.excluded_calibration_m = float()
        self.calibration_m = float()
        self.calibration_b = float()
        self.calibration_r_squared = float()
        self.calibration_times = np.array([])
        self.calibration_volumes = np.array([])
        self.calibration_pressures = np.array([])
        self.excluded_calibration_times = np.array([])
        self.excluded_calibration_volumes = np.array([])
        self.excluded_calibration_pressures = np.array([])
        self.landmarks = []
        self.stats = []
        self.avg_stats = []
        self.stdev_stats = []

    def add_calibration_point(self, time, pressure, volume):
        self.calibration_times = np.append(self.calibration_times, time)
        self.calibration_pressures = np.append(self.calibration_pressures, pressure)
        self.calibration_volumes = np.append(self.calibration_volumes, volume)

    def add_excluded_calibration_point(self, time, pressure, volume):
        self.excluded_calibration_times = np.append(self.excluded_calibration_times, time)
        self.excluded_calibration_pressures = np.append(self.excluded_calibration_pressures, pressure)
        self.excluded_calibration_volumes = np.append(self.excluded_calibration_volumes, volume)

    def calculate_breath_stats(self):
        self.start_idx = np.ma.asarray(np.full(self.num_breaths, np.nan, dtype=int))
        self.stop_idx = np.ma.asarray(np.full(self.num_breaths, np.nan, dtype=int))
        self.i50v = np.ma.asarray(np.full(self.num_breaths, np.nan, dtype=int))
        self.i64v = np.ma.asarray(np.full(self.num_breaths, np.nan, dtype=int))
        self.e50v = np.ma.asarray(np.full(self.num_breaths, np.nan, dtype=int))
        self.e64v = np.ma.asarray(np.full(self.num_breaths, np.nan, dtype=int))

        self.landmarks = [self.inhale_flow_peaks,
                          self.exhale_flow_peaks,
                          self.inhale_onsets,
                          self.inhale_offsets,
                          self.inhale_pause_onsets,
                          self.inhale_pause_offsets,
                          self.exhale_onsets,
                          self.exhale_offsets,
                          self.exhale_pause_onsets,
                          self.exhale_pause_offsets,
                          self.inhale_volume_peaks,
                          self.exhale_volume_peaks,
                          self.start_idx,
                          self.stop_idx,
                          self.i50v,
                          self.i64v,
                          self.e50v,
                          self.e64v]

        self.inhale_volumes = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.inhale_pause_volumes = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.exhale_volumes = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.exhale_pause_volumes = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.inhale_pause_plus_exhale_volumes = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.inhale_peak_flows = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.exhale_peak_flows = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.if50 = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.ef50 = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.inhale_times = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.inhale_pause_times = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.exhale_times = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.exhale_pause_times = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.inhale_pause_plus_exhale_times = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.breath_times = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.bpm = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.tri_50 = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.tri_64 = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.tre_50 = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.tre_64 = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.penh = np.ma.asarray(np.full(self.num_breaths, np.nan))
        self.saturated = np.full(self.num_breaths, np.nan)
        self.movement = np.full(self.num_breaths, np.nan)
        self.outlier = np.full(self.num_breaths, np.nan)

        self.stats = [self.inhale_volumes,
                      self.inhale_pause_volumes,
                      self.exhale_volumes,
                      self.exhale_pause_volumes,
                      self.inhale_pause_plus_exhale_volumes,
                      self.inhale_peak_flows,
                      self.exhale_peak_flows,
                      self.if50,
                      self.ef50,
                      self.inhale_times,
                      self.inhale_pause_times,
                      self.exhale_times,
                      self.exhale_pause_times,
                      self.inhale_pause_plus_exhale_times,
                      self.breath_times,
                      self.bpm,
                      self.tri_50,
                      self.tri_64,
                      self.tre_50,
                      self.tre_64,
                      self.penh]

        for idx in range(self.num_breaths):
            # landmarks, saved as index where landmark occurs within respiration data
            self.start_idx[idx] = int(self.inhale_onsets[idx])
            if self.exhale_pause_offsets[idx] >= 0:
                self.stop_idx[idx] = self.exhale_pause_offsets[idx]
            else:
                self.stop_idx[idx] = self.exhale_offsets[idx]
            self.i50v[idx] = self.find_relaxation_idx(self.local_corrected_volume, self.inhale_onsets[idx],
                                                      self.inhale_offsets[idx], 50)
            self.i64v[idx] = self.find_relaxation_idx(self.local_corrected_volume, self.inhale_onsets[idx],
                                                      self.inhale_offsets[idx], 64)
            self.e50v[idx] = self.find_relaxation_idx(self.local_corrected_volume, self.exhale_onsets[idx],
                                                      self.exhale_offsets[idx], 50)
            self.e64v[idx] = self.find_relaxation_idx(self.local_corrected_volume, self.exhale_onsets[idx],
                                                      self.exhale_offsets[idx], 64)

            # Volume stats
            self.inhale_volumes[idx] = (self.local_corrected_volume[self.inhale_offsets[idx]] -
                                        self.local_corrected_volume[self.inhale_onsets[idx]])
            self.inhale_pause_volumes[idx] = (self.local_corrected_volume[self.inhale_pause_offsets[idx]] -
                                              self.local_corrected_volume[self.inhale_pause_onsets[idx]])
            self.exhale_volumes[idx] = (self.local_corrected_volume[self.exhale_offsets[idx]] -
                                        self.local_corrected_volume[self.exhale_onsets[idx]])
            self.exhale_pause_volumes[idx] = (self.local_corrected_volume[self.exhale_pause_offsets[idx]] -
                                              self.local_corrected_volume[self.exhale_pause_onsets[idx]])
            self.inhale_pause_plus_exhale_volumes[idx] = self.inhale_pause_volumes[idx] + self.exhale_volumes[idx]

            # Flow stats
            self.inhale_peak_flows[idx] = self.flow[self.inhale_flow_peaks[idx]]
            self.exhale_peak_flows[idx] = self.flow[self.exhale_flow_peaks[idx]]

            # Mid-inspiratory flow
            self.if50[idx] = self.flow[self.i50v[idx]]

            # Mid-expiratory flow
            self.ef50[idx] = self.flow[self.e50v[idx]]

            # Time stats
            self.inhale_times[idx] = (self.time[self.inhale_offsets[idx]] -
                                      self.time[self.inhale_onsets[idx]])
            self.inhale_pause_times[idx] = (self.time[self.inhale_pause_offsets[idx]] -
                                            self.time[self.inhale_pause_onsets[idx]])
            self.exhale_times[idx] = (self.time[self.exhale_offsets[idx]] -
                                      self.time[self.exhale_onsets[idx]])
            self.exhale_pause_times[idx] = (self.time[self.exhale_pause_offsets[idx]] -
                                            self.time[self.exhale_pause_onsets[idx]])
            self.inhale_pause_plus_exhale_times[idx] = self.inhale_pause_times[idx] + self.exhale_times[idx]

            self.breath_times[idx] = (self.time[self.stop_idx[idx]] -
                                      self.time[self.start_idx[idx]])
            self.bpm[idx] = 60.0 / self.breath_times[idx]

            # Relaxation times to 50% or 64% inhale/exhale volume
            self.tri_50[idx] = self.time[self.i50v[idx]] - self.time[self.inhale_onsets[idx]]
            self.tri_64[idx] = self.time[self.i64v[idx]] - self.time[self.inhale_onsets[idx]]
            self.tre_50[idx] = self.time[self.e50v[idx]] - self.time[self.exhale_onsets[idx]]
            self.tre_64[idx] = self.time[self.e64v[idx]] - self.time[self.exhale_onsets[idx]]

            # Enhanced pause
            self.penh[idx] = ((self.exhale_times[idx] / self.tre_64[idx] - 1) *
                              abs(self.exhale_peak_flows[idx] / self.inhale_peak_flows[idx]))

    def calculate_flow_signal(self, data):
        # Calculate flow from global drift corrected volume using cubic spline interpolation
        cs = interpolate.CubicSpline(self.time, data)
        self.flow = np.array(cs(self.time, 1))

    def calculate_pressure_signal(self, supply_voltage, min_pascals, max_pascals):
        pressure = voltage_to_pressure(self.voltage, supply_voltage, min_pascals, max_pascals)
        self.pressure = pressure

    def calculate_volume_signal(self, ref_pressure):
        volume = pressure_to_volume(self.pressure, ref_pressure, self.calibration_m, self.calibration_b)
        # TODO: Should self.volume be a tuple? Can this be removed without wrecking program?
        self.volume = tuple(volume)
        if self.lowpass_pressure.size == 0:
            # Lowpass pressure signal was never calculated, fill array with NaN values for saving data later
            self.lowpass_pressure = np.full(self.len_data, np.nan)

    def calculate_window_avgs(self, window_size, window_offset, roi_start_idx, roi_stop_idx):
        window_size = round(window_size * self.sampling_rate)
        window_offset = round(window_offset * self.sampling_rate)
        window_start_idx = roi_start_idx
        window_stop_idx = window_start_idx + window_size
        len_roi = roi_stop_idx - roi_start_idx
        if len_roi < window_size:
            # want at least one window, even if it's not full window_size
            num_windows = 1
        else:
            # otherwise, each window must be full window_size; remainder data points won't be averaged
            num_windows = int(np.floor((len_roi - window_size) / window_offset) + 1)
        self.num_windows = num_windows

        # Setup stat average and standard deviation arrays
        self.avg_inhale_volume = np.full(num_windows, np.nan)
        self.avg_inhale_pause_volume = np.full(num_windows, np.nan)
        self.avg_exhale_volume = np.full(num_windows, np.nan)
        self.avg_exhale_pause_volume = np.full(num_windows, np.nan)
        self.avg_inhale_pause_plus_exhale_volume = np.full(num_windows, np.nan)
        self.avg_inhale_peak_flow = np.full(num_windows, np.nan)
        self.avg_exhale_peak_flow = np.full(num_windows, np.nan)
        self.avg_if50 = np.full(num_windows, np.nan)
        self.avg_ef50 = np.full(num_windows, np.nan)
        self.avg_inhale_time = np.full(num_windows, np.nan)
        self.avg_inhale_pause_time = np.full(num_windows, np.nan)
        self.avg_exhale_time = np.full(num_windows, np.nan)
        self.avg_exhale_pause_time = np.full(num_windows, np.nan)
        self.avg_inhale_pause_plus_exhale_time = np.full(num_windows, np.nan)
        self.avg_breath_time = np.full(num_windows, np.nan)
        self.avg_bpm = np.full(num_windows, np.nan)
        self.avg_tri_50 = np.full(num_windows, np.nan)
        self.avg_tri_64 = np.full(num_windows, np.nan)
        self.avg_tre_50 = np.full(num_windows, np.nan)
        self.avg_tre_64 = np.full(num_windows, np.nan)
        self.avg_penh = np.full(num_windows, np.nan)

        self.avg_stats = [self.avg_inhale_volume,
                          self.avg_inhale_pause_volume,
                          self.avg_exhale_volume,
                          self.avg_exhale_pause_volume,
                          self.avg_inhale_pause_plus_exhale_volume,
                          self.avg_inhale_peak_flow,
                          self.avg_exhale_peak_flow,
                          self.avg_if50,
                          self.avg_ef50,
                          self.avg_inhale_time,
                          self.avg_inhale_pause_time,
                          self.avg_exhale_time,
                          self.avg_exhale_pause_time,
                          self.avg_inhale_pause_plus_exhale_time,
                          self.avg_breath_time,
                          self.avg_bpm,
                          self.avg_tri_50,
                          self.avg_tri_64,
                          self.avg_tre_50,
                          self.avg_tre_64,
                          self.avg_penh]

        self.stdev_inhale_volume = np.full(num_windows, np.nan)
        self.stdev_inhale_pause_volume = np.full(num_windows, np.nan)
        self.stdev_exhale_volume = np.full(num_windows, np.nan)
        self.stdev_exhale_pause_volume = np.full(num_windows, np.nan)
        self.stdev_inhale_pause_plus_exhale_volume = np.full(num_windows, np.nan)
        self.stdev_inhale_peak_flow = np.full(num_windows, np.nan)
        self.stdev_exhale_peak_flow = np.full(num_windows, np.nan)
        self.stdev_if50 = np.full(num_windows, np.nan)
        self.stdev_ef50 = np.full(num_windows, np.nan)
        self.stdev_inhale_time = np.full(num_windows, np.nan)
        self.stdev_inhale_pause_time = np.full(num_windows, np.nan)
        self.stdev_exhale_time = np.full(num_windows, np.nan)
        self.stdev_exhale_pause_time = np.full(num_windows, np.nan)
        self.stdev_inhale_pause_plus_exhale_time = np.full(num_windows, np.nan)
        self.stdev_breath_time = np.full(num_windows, np.nan)
        self.stdev_bpm = np.full(num_windows, np.nan)
        self.stdev_tri_50 = np.full(num_windows, np.nan)
        self.stdev_tri_64 = np.full(num_windows, np.nan)
        self.stdev_tre_50 = np.full(num_windows, np.nan)
        self.stdev_tre_64 = np.full(num_windows, np.nan)
        self.stdev_penh = np.full(num_windows, np.nan)

        self.stdev_stats = [self.stdev_inhale_volume,
                            self.stdev_inhale_pause_volume,
                            self.stdev_exhale_volume,
                            self.stdev_exhale_pause_volume,
                            self.stdev_inhale_pause_plus_exhale_volume,
                            self.stdev_inhale_peak_flow,
                            self.stdev_exhale_peak_flow,
                            self.stdev_if50,
                            self.stdev_ef50,
                            self.stdev_inhale_time,
                            self.stdev_inhale_pause_time,
                            self.stdev_exhale_time,
                            self.stdev_exhale_pause_time,
                            self.stdev_inhale_pause_plus_exhale_time,
                            self.stdev_breath_time,
                            self.stdev_bpm,
                            self.stdev_tri_50,
                            self.stdev_tri_64,
                            self.stdev_tre_50,
                            self.stdev_tre_64,
                            self.stdev_penh]

        self.window_start_time = np.asarray(np.full(num_windows, np.nan))
        self.window_stop_time = np.asarray(np.full(num_windows, np.nan))
        self.window_rel_center_time = np.asarray(np.full(num_windows, np.nan))
        self.window_num_breaths = np.asarray(np.full(num_windows, np.nan))
        self.window_num_excluded = np.asarray(np.full(num_windows, np.nan))
        self.window_num_saturated = np.asarray(np.full(num_windows, np.nan))
        self.window_num_movement = np.asarray(np.full(num_windows, np.nan))
        self.window_num_outlier = np.asarray(np.full(num_windows, np.nan))

        for window_idx in range(num_windows):
            # determine which breaths fall within window - any breaths with start_idx in window
            data_start_idx = np.ma.argmax(self.start_idx >= window_start_idx)
            if self.start_idx[-1] < window_stop_idx:
                data_stop_idx = len(self.start_idx)
            else:
                data_stop_idx = np.ma.argmin(self.start_idx < window_stop_idx)

            self.window_start_time[window_idx] = self.time[window_start_idx]
            if len_roi < window_size:
                self.window_stop_time[window_idx] = self.time[roi_stop_idx]
                self.window_rel_center_time[window_idx] = round((len_roi / 2.0) / self.sampling_rate, 3)
            else:
                self.window_stop_time[window_idx] = self.time[window_stop_idx]
                self.window_rel_center_time[window_idx] = round(((window_size / 2.0) + window_idx * window_offset) /
                                                                self.sampling_rate, 3)
            self.window_num_breaths[window_idx] = self.inhale_volumes[data_start_idx:data_stop_idx].size
            self.window_num_excluded[window_idx] = np.sum(self.breaths_to_exclude[data_start_idx:data_stop_idx])
            self.window_num_saturated[window_idx] = np.sum(self.saturated[data_start_idx:data_stop_idx])
            self.window_num_movement[window_idx] = np.sum(self.movement[data_start_idx:data_stop_idx])
            self.window_num_outlier[window_idx] = np.sum(self.outlier[data_start_idx:data_stop_idx])

            for stat_idx in range(len(self.stats)):
                stat = self.stats[stat_idx]
                if stat[data_start_idx:data_stop_idx].count() == 0:
                    # All data has been masked/excluded
                    self.avg_stats[stat_idx][window_idx] = np.nan
                    self.stdev_stats[stat_idx][window_idx] = np.nan
                else:
                    # At least one data point has not been excluded, np.mean should return a float
                    self.avg_stats[stat_idx][window_idx] = np.mean(stat[data_start_idx:data_stop_idx])
                    self.stdev_stats[stat_idx][window_idx] = np.std(stat[data_start_idx:data_stop_idx])

            # shift window by window_offset
            window_start_idx = window_start_idx + window_offset
            window_stop_idx = window_stop_idx + window_offset

    def determine_breaths_to_exclude(self, adjacency, must_be_between_exclusions, remove_saturated, remove_movement,
                                     remove_outlier):
        self.breaths_to_exclude = np.full(self.num_breaths, False)
        for idx in range(self.num_breaths):
            if ((self.saturated[idx] and remove_saturated) or
                    (self.movement[idx] and remove_movement) or
                    (self.outlier[idx] and remove_outlier)):
                # Do not analyze this breath
                self.breaths_to_exclude[idx] = True

        self.exclude_breaths_by_proximity(adjacency, must_be_between_exclusions)

        for landmark in self.landmarks:
            landmark.mask = self.breaths_to_exclude

        for stat in self.stats:
            stat.mask = self.breaths_to_exclude

    @staticmethod
    def determine_pause_search_range(volume_data, previous_troughs, current_peaks, current_troughs,
                                     min_inhale_vol_idx, max_inhale_exhale_vol_idx, min_exhale_vol_idx,
                                     pause_search_threshold):
        # Determine threshold volumes
        min_inhale_volume = volume_data[min_inhale_vol_idx]
        max_inhale_exhale_volume = volume_data[max_inhale_exhale_vol_idx]
        min_exhale_volume = volume_data[min_exhale_vol_idx]
        inhale_window = volume_data[min_inhale_vol_idx:max_inhale_exhale_vol_idx]
        exhale_window = volume_data[max_inhale_exhale_vol_idx:min_exhale_vol_idx]
        inhale_volume = max_inhale_exhale_volume - min_inhale_volume
        exhale_volume = max_inhale_exhale_volume - min_exhale_volume
        inhale_vol_threshold = min_inhale_volume + (pause_search_threshold / 100.0) * inhale_volume
        exhale_vol_threshold = max_inhale_exhale_volume - (pause_search_threshold / 100.0) * exhale_volume

        # Find inhale/exhale indices where threshold is first exceeded
        inhale_vol_above_threshold = np.where(inhale_window > inhale_vol_threshold)
        exhale_vol_below_threshold = np.where(exhale_window < exhale_vol_threshold)
        try:
            inhale_threshold_idx = inhale_vol_above_threshold[0][0] + min_inhale_vol_idx
        except IndexError:
            inhale_threshold_idx = max_inhale_exhale_vol_idx - 1
        try:
            exhale_threshold_idx = exhale_vol_below_threshold[0][0] + max_inhale_exhale_vol_idx
        except IndexError:
            exhale_threshold_idx = min_exhale_vol_idx - 1

        troubleshooting = False
        if troubleshooting:
            print('min_inhale_vol: %.3f, max_inhale_exhale_vol: %.3f, min_exhale_vol: %.3f' % (min_inhale_volume,
                                                                                               max_inhale_exhale_volume,
                                                                                               min_exhale_volume))
            print('inhale_vol: %.3f, exhale_vol: %.3f' % (inhale_volume, exhale_volume))
            print('inhale_vol_threshold: %.3f, exhale_vol_threshold: %.3f' % (inhale_vol_threshold,
                                                                              exhale_vol_threshold))
            print('inhale_threshold_idx: %.3f, exhale_threshold_idx: %.3f' % (inhale_threshold_idx,
                                                                              exhale_threshold_idx))
            print('thresh_inhale_idx_val: %.3f, thresh_exhale_idx_val: %.3f' % (volume_data[inhale_threshold_idx],
                                                                                volume_data[exhale_threshold_idx]))
            print('prev_inhale_idx_val: %.3f, prev_exhale_idx_val: %.3f' % (volume_data[inhale_threshold_idx - 1],
                                                                            volume_data[exhale_threshold_idx - 1]))
            print()

        vol_threshold_idxs = tuple([inhale_threshold_idx, exhale_threshold_idx])

        # End exhale pause search starts at last trough of previous breath
        eep_search_start_idx = int(previous_troughs[-1])

        # End exhale pause search stops at first peak of current breath that exceeds pause_search_threshold
        current_peak_values = np.array([volume_data[int(idx)] for idx in current_peaks])
        try:
            peak_to_use = np.where(current_peak_values > inhale_vol_threshold)[0][0]  # first exceeding peak
            # peak_to_use = np.where(current_peak_values <= inhale_vol_threshold)[0][-1]  # last non-exceeding peak
        except IndexError:
            peak_to_use = len(current_peak_values) - 1
            # peak_to_use = 0
        eep_search_stop_idx = int(current_peaks[peak_to_use])

        # End inhale pause search starts at last peak of current breath
        eip_search_start_idx = int(current_peaks[-1])

        # End inhale pause search stops at first trough of current breath that exceeds pause_search_threshold
        current_trough_values = np.array([volume_data[int(idx)] for idx in current_troughs])
        try:
            trough_to_use = np.where(current_trough_values < exhale_vol_threshold)[0][0]  # first exceeding trough
            # trough_to_use = np.where(current_trough_values >= exhale_vol_threshold)[0][-1] # last non-exceeding trough
        except IndexError:
            trough_to_use = len(current_trough_values) - 1
            # trough_to_use = 0
        eip_search_stop_idx = int(current_troughs[trough_to_use])

        search_idxs = tuple([eep_search_start_idx, eep_search_stop_idx, eip_search_start_idx, eip_search_stop_idx])
        return search_idxs, vol_threshold_idxs

    def exclude_breaths_by_proximity(self, adjacency, must_be_between_exclusions=True):
        """
        This method finds and excludes breaths that are within adjacency breaths away from an excluded breath. If
        must_be_between_exclusions is True, then breaths can only be excluded if this condtion applies in both
        directions.
        :param adjacency: distance from an excluded breath required to exclude by proximity
        :param must_be_between_exclusions: flag that limits exclusion by proximity, requiring bi-directional adjacent
        exclusions
        :return:
        """

        self.proximity = np.full(self.num_breaths, False)
        already_excluded = self.breaths_to_exclude.copy()
        for idx in range(self.num_breaths):
            # Skip breaths that have already been excluded
            if already_excluded[idx]:
                continue

            exclude_by_proximity = False

            # Handle edge cases
            if idx < adjacency:
                left_idx = 0
            else:
                left_idx = idx - adjacency
            if idx > self.num_breaths - adjacency:
                right_idx = self.num_breaths
            else:
                right_idx = idx + 1 + adjacency

            # Search for adjacent exclusions
            outlier_to_left = already_excluded[left_idx:idx].any()
            outlier_to_right = already_excluded[idx + 1:right_idx].any()

            # Exclude currenty breath if appropriate
            if outlier_to_left and outlier_to_right:
                exclude_by_proximity = True
            elif (outlier_to_left or outlier_to_right) and not must_be_between_exclusions:
                exclude_by_proximity = True

            self.proximity[idx] = exclude_by_proximity
            self.breaths_to_exclude[idx] = exclude_by_proximity

    def find_extrema(self, data, roi_start_idx, roi_stop_idx, window_sizes, shifts):
        peak_threshold = np.mean(data) + 0.5 * np.std(data)
        trough_threshold = np.mean(data) - 0.5 * np.std(data)

        # Pad data on both ends with zeros the size of the largest window, so every true data point will be examined
        # by each window_size * shift iteration. Otherwise, peaks/troughs at edges of data could be missed. "Zero" is
        # defined by the mean of peak_threshold and trough_thresholds, so no extrema will ever be detected in padding.
        padding_length = max(window_sizes)
        padding = np.array([np.average([peak_threshold, trough_threshold])] * padding_length)
        padded_data = np.concatenate((padding, data, padding))

        num_iterations = len(window_sizes) * len(shifts)
        iteration = 0
        num_peaks_by_idx = np.zeros(self.len_data, dtype=int)
        num_troughs_by_idx = np.zeros(self.len_data, dtype=int)

        for idx_window_sizes, window_size in enumerate(window_sizes):
            for idx_shifts, shift in enumerate(shifts):

                start_idx = roi_start_idx + int(round(shift * window_size))
                stop_idx = start_idx + window_size
                while start_idx < roi_stop_idx + 2 * padding_length:
                    if stop_idx > roi_stop_idx + 2 * padding_length:
                        window = padded_data[start_idx::]
                    else:
                        window = padded_data[start_idx:stop_idx]

                    # identify if window max is a peak
                    window_max_idx = np.argmax(window)  # index within window
                    window_max = window[window_max_idx]
                    if window_max >= peak_threshold:
                        # Ensure max index does not lie within padding
                        window_max_idx = window_max_idx + start_idx - padding_length  # index within unpadded data
                        if roi_start_idx <= window_max_idx < roi_stop_idx:
                            num_peaks_by_idx[window_max_idx] = num_peaks_by_idx[window_max_idx] + 1
                        else:
                            # print('Peak index lies in padding')
                            pass

                    # identify if window min is a trough
                    window_min_idx = np.argmin(window)  # index within window
                    window_min = window[window_min_idx]
                    if window_min <= trough_threshold:
                        # Ensure min index does not lie within padding
                        window_min_idx = window_min_idx + start_idx - padding_length  # index within unpadded data
                        if 0 <= window_min_idx < self.len_data:
                            num_troughs_by_idx[window_min_idx] = num_troughs_by_idx[window_min_idx] + 1
                        else:
                            # print('Trough index lies in padding')
                            pass

                    # print(window, min(window), max(window))
                    start_idx = stop_idx
                    stop_idx = start_idx + window_size

                iteration += 1

        num_peaks_found = np.zeros(num_iterations, dtype=int)
        num_troughs_found = np.zeros(num_iterations, dtype=int)

        for idx in range(num_iterations):
            num_peaks_found[idx] = (num_peaks_by_idx > idx).sum()
            num_troughs_found[idx] = (num_troughs_by_idx > idx).sum()

        peak_diff = np.diff(num_peaks_found)
        trough_diff = np.diff(num_troughs_found)

        best_peak_consensus_threshold = np.argmax(peak_diff)
        best_trough_consensus_threshold = np.argmax(trough_diff)

        best_consensus_threshold = int(round(np.average([best_peak_consensus_threshold,
                                                         best_trough_consensus_threshold])))

        consensus_peaks = np.ma.asarray(num_peaks_by_idx > best_consensus_threshold).nonzero()[0]
        consensus_troughs = np.ma.asarray(num_troughs_by_idx > best_consensus_threshold).nonzero()[0]

        # Peaks and troughs should alternate. If consecutive peaks/troughs, algorithm will handle later.
        # For ease of rest of algorithm, begin with a trough (exhalation) and end with a peak (inhalation).
        consecutive_peak_idx = np.array([], dtype=int)
        consecutive_trough_idx = np.array([], dtype=int)
        all_peaks = []
        all_troughs = []
        max_peaks = np.ma.array([], dtype=int)
        min_troughs = np.ma.array([], dtype=int)
        min_inhale_vol_idx = np.array([], dtype=int)
        max_inhale_exhale_vol_idx = np.array([], dtype=int)
        min_exhale_vol_idx = np.array([], dtype=int)

        # Find first trough and first peak occurring after that trough
        troughs_list_idx = 0
        peaks_list_idx = int(np.argmax(consensus_peaks > consensus_troughs[troughs_list_idx]))

        # Pad end of troughs list with dummy trough if last extrema is a peak
        if consensus_peaks[-1] > consensus_troughs[-1]:
            consensus_troughs = np.append(consensus_troughs, int(consensus_peaks[-1] + 1))

        # Find alternating peaks and troughs
        while (peaks_list_idx < len(consensus_peaks)) and (troughs_list_idx < len(consensus_troughs)):
            current_peak_idx = int(consensus_peaks[peaks_list_idx])
            current_trough_idx = int(consensus_troughs[troughs_list_idx])

            if current_peak_idx < current_trough_idx:
                # save previous trough(s)
                if consecutive_trough_idx.size:
                    all_troughs.append(consecutive_trough_idx)

                    # Find and save min flow trough
                    consecutive_trough_values = np.array([data[int(idx)] for idx in consecutive_trough_idx])
                    min_trough_idx = int(np.argmin(consecutive_trough_values))
                    min_troughs = np.append(min_troughs, min_trough_idx)

                    # Find and save min volume idx between last trough and current peak
                    last_trough_idx = int(consecutive_trough_idx[-1])
                    rel_min_vol_idx = int(np.argmin(self.local_corrected_volume[last_trough_idx:current_peak_idx + 1]))
                    abs_min_vol_idx = last_trough_idx + rel_min_vol_idx
                    if min_inhale_vol_idx.size:
                        # Add min volume index to previous breath, except for 1st breath
                        min_exhale_vol_idx = np.append(min_exhale_vol_idx, abs_min_vol_idx)
                    min_inhale_vol_idx = np.append(min_inhale_vol_idx, abs_min_vol_idx)

                    # Reset consecutive troughs array
                    consecutive_trough_idx = np.array([])

                # keep all consecutive peaks found
                consecutive_peak_idx = np.append(consecutive_peak_idx, current_peak_idx)

                # advance to next peak index
                peaks_list_idx = peaks_list_idx + 1

            else:
                # save previous peak(s)
                if consecutive_peak_idx.size:
                    all_peaks.append(consecutive_peak_idx)

                    # Find and save max flow peak
                    consecutive_peak_values = np.array([data[int(idx)] for idx in consecutive_peak_idx])
                    max_peak_idx = int(np.argmax(consecutive_peak_values))
                    max_peaks = np.append(max_peaks, max_peak_idx)

                    # Find and save max volume idx between last peak and current trough
                    last_peak_idx = int(consecutive_peak_idx[-1])
                    rel_max_vol_idx = int(np.argmax(self.local_corrected_volume[last_peak_idx:current_trough_idx + 1]))
                    abs_max_vol_idx = last_peak_idx + rel_max_vol_idx
                    max_inhale_exhale_vol_idx = np.append(max_inhale_exhale_vol_idx, abs_max_vol_idx)

                    # Reset consecutive peaks array
                    consecutive_peak_idx = np.array([])

                # keep all consecutive troughs found
                consecutive_trough_idx = np.append(consecutive_trough_idx, current_trough_idx)

                # advance to next trough index
                troughs_list_idx = troughs_list_idx + 1

        # The initial trough and final peak will not be included in final breathing data, but will be useful for next
        # steps of algorithm. Separate them here.
        self.final_peak = consecutive_peak_idx
        min_inhale_vol_idx = np.delete(min_inhale_vol_idx, -1)

        self.min_inhale_volume_idx = min_inhale_vol_idx
        self.max_inhale_exhale_volume_idx = max_inhale_exhale_vol_idx
        self.min_exhale_volume_idx = min_exhale_vol_idx

        self.consensus_all_peaks = all_peaks
        self.initial_trough = all_troughs[0]
        self.consensus_all_troughs = all_troughs[1::]
        self.num_breaths = len(all_peaks)

        # Flow min/max for each breath
        self.consensus_peaks = max_peaks
        self.consensus_troughs = np.delete(min_troughs, 0)

    def find_landmarks(self, flow_data, volume_data, pause_search_threshold, n_bins, extra_zero_bin_threshold,
                       extra_pause_bin_threshold, mode_bin_ratio, extra_pause_bin_ratio):
        # Zero-cross threshold, where "zero" is the mean of data
        zero_cross_threshold = np.mean(flow_data)
        self.data_zero = zero_cross_threshold

        inhale_onsets = np.ma.asarray(np.array([], dtype=int))
        exhale_pause_onsets = np.ma.asarray(np.array([], dtype=int))
        exhale_onsets = np.ma.asarray(np.array([], dtype=int))
        inhale_pause_onsets = np.ma.asarray(np.array([], dtype=int))
        final_inhale_onset = -1
        num_breaths = self.num_breaths

        for idx in range(num_breaths + 1):

            # handle first and last inhale onsets differently
            if idx == 0:
                min_inhale_vol_idx = self.min_inhale_volume_idx[idx]
                max_inhale_exhale_vol_idx = self.max_inhale_exhale_volume_idx[idx]
                min_exhale_vol_idx = self.min_exhale_volume_idx[idx]
                previous_troughs = self.initial_trough
                current_peaks = self.consensus_all_peaks[idx]
                current_troughs = self.consensus_all_troughs[idx]

            elif idx == num_breaths:
                min_inhale_vol_idx = self.min_exhale_volume_idx[idx - 1]
                max_inhale_exhale_vol_idx = self.max_inhale_exhale_volume_idx[idx - 1]  # use last breath exhale volume
                min_exhale_vol_idx = max_inhale_exhale_vol_idx + 1  # dummy index, this won't be used
                previous_troughs = self.consensus_all_troughs[idx - 1]
                current_peaks = self.final_peak
                current_troughs = [self.final_peak[-1] + 1]  # dummy index, this won't be used

            else:
                min_inhale_vol_idx = self.min_inhale_volume_idx[idx]
                max_inhale_exhale_vol_idx = self.max_inhale_exhale_volume_idx[idx]
                min_exhale_vol_idx = self.min_exhale_volume_idx[idx]
                previous_troughs = self.consensus_all_troughs[idx - 1]
                current_peaks = self.consensus_all_peaks[idx]
                current_troughs = self.consensus_all_troughs[idx]

            # Determine pause search start/stop indices (eep = end exhale pause, eip = end inhale pause)
            search_idxs, vol_threshold_idxs = self.determine_pause_search_range(volume_data, previous_troughs,
                                                                                current_peaks, current_troughs,
                                                                                min_inhale_vol_idx,
                                                                                max_inhale_exhale_vol_idx,
                                                                                min_exhale_vol_idx,
                                                                                pause_search_threshold)
            eep_search_start_idx = search_idxs[0]
            eep_search_stop_idx = search_idxs[1]
            eip_search_start_idx = search_idxs[2]
            eip_search_stop_idx = search_idxs[3]
            inhale_vol_threshold_idx = vol_threshold_idxs[0]
            exhale_vol_threshold_idx = vol_threshold_idxs[1]

            # Not looking for exhale pauses; if desired, can toggle find_pause to True
            # Create histogram from last trough to current peak to find end exhale pause and inhale onsets
            inhale_onset, last_exhale_pause_onset = get_onsets_from_hist(flow_data, volume_data, eep_search_start_idx,
                                                                         eep_search_stop_idx, inhale_vol_threshold_idx,
                                                                         True, zero_cross_threshold, n_bins,
                                                                         extra_zero_bin_threshold,
                                                                         extra_pause_bin_threshold,
                                                                         mode_bin_ratio, extra_pause_bin_ratio,
                                                                         find_pause=False, show_hist=False)
            if idx < num_breaths:
                inhale_onsets = np.append(inhale_onsets, inhale_onset)
            elif idx == num_breaths:
                final_inhale_onset = inhale_onset
            if idx != 0:
                exhale_pause_onsets = np.append(exhale_pause_onsets, last_exhale_pause_onset)

            # Create histogram from current peak to current trough to find end inhale pause and exhale onsets
            if idx < num_breaths:
                exhale_onset, inhale_pause_onset = get_onsets_from_hist(flow_data, volume_data, eip_search_start_idx,
                                                                        eip_search_stop_idx, exhale_vol_threshold_idx,
                                                                        False, zero_cross_threshold, n_bins,
                                                                        extra_zero_bin_threshold,
                                                                        extra_pause_bin_threshold,
                                                                        mode_bin_ratio, extra_pause_bin_ratio,
                                                                        find_pause=True, show_hist=False)

                exhale_onsets = np.append(exhale_onsets, exhale_onset)
                inhale_pause_onsets = np.append(inhale_pause_onsets, inhale_pause_onset)

        # Find offset indices
        inhale_offsets = np.ma.asarray(np.full(num_breaths, -1, dtype=int))
        inhale_pause_offsets = np.ma.asarray(np.full(num_breaths, -1, dtype=int))
        inhale_volume_peaks = np.ma.asarray(np.full(num_breaths, -1, dtype=int))
        inhale_flow_peaks = np.ma.asarray(np.full(num_breaths, -1, dtype=int))
        exhale_offsets = np.ma.asarray(np.full(num_breaths, -1, dtype=int))
        exhale_pause_offsets = np.ma.asarray(np.full(num_breaths, -1, dtype=int))
        exhale_volume_peaks = np.ma.asarray(np.full(num_breaths, -1, dtype=int))
        exhale_flow_peaks = np.ma.asarray(np.full(num_breaths, -1, dtype=int))

        for idx in range(num_breaths):
            # need to get last breath's exhale offset differently
            if (idx == num_breaths - 1) and (final_inhale_onset != -1):
                next_inhale_onset = final_inhale_onset
            else:
                next_inhale_onset = inhale_onsets[idx + 1]
            inhale_onset = inhale_onsets[idx]
            inhale_pause_onset = inhale_pause_onsets[idx]
            exhale_onset = exhale_onsets[idx]
            exhale_pause_onset = exhale_pause_onsets[idx]

            # Inhale offsets
            if inhale_pause_onset >= 0:
                inhale_offsets[idx] = inhale_pause_onset
                inhale_pause_offsets[idx] = exhale_onset
            else:
                inhale_offsets[idx] = exhale_onset
                inhale_pause_offsets[idx] = -1
            inhale_volume_peaks[idx] = int(np.argmax(self.local_corrected_volume[inhale_onset:exhale_onset])
                                           + inhale_onset)
            inhale_flow_peaks[idx] = int(np.argmax(self.flow[inhale_onset:exhale_onset]) + inhale_onset)

            # Exhale offsets
            if exhale_pause_onset >= 0:
                exhale_offsets[idx] = exhale_pause_onset
                exhale_pause_offsets[idx] = next_inhale_onset
            else:
                exhale_offsets[idx] = next_inhale_onset
                exhale_pause_onsets[idx] = -1
            exhale_volume_peaks[idx] = int(np.argmin(self.local_corrected_volume[exhale_onset:next_inhale_onset])
                                           + exhale_onset)
            exhale_flow_peaks[idx] = int(np.argmin(self.flow[exhale_onset:next_inhale_onset]) + exhale_onset)

        self.num_breaths = num_breaths
        self.inhale_onsets = inhale_onsets
        self.inhale_offsets = inhale_offsets
        self.inhale_pause_onsets = inhale_pause_onsets
        self.inhale_pause_offsets = inhale_pause_offsets
        self.exhale_onsets = exhale_onsets
        self.exhale_offsets = exhale_offsets
        self.exhale_pause_onsets = exhale_pause_onsets
        self.exhale_pause_offsets = exhale_pause_offsets
        self.inhale_volume_peaks = inhale_volume_peaks
        self.exhale_volume_peaks = exhale_volume_peaks
        self.inhale_flow_peaks = inhale_flow_peaks
        self.exhale_flow_peaks = exhale_flow_peaks

    def find_movement_breaths_testing(self, movement_threshold=0.5, remove_movement=True):
        """
        Method for detecting breaths where movement has occured. Look at previous breath; if peaks or troughs have
        changed by more than the movement_threshold for the current breath, movement is detected.
        :param movement_threshold:
        :param remove_movement:
        :return:
        """
        self.movement = np.full(self.num_breaths, False)

        if not remove_movement or self.num_breaths < 2:
            return

        # Set threshold at percentage of previous breath's tidal volume
        threshold = movement_threshold * np.mean([abs(self.inhale_volumes[0]), abs(self.exhale_volumes[0])])

        prev_inhale_onset = self.local_corrected_volume[self.inhale_onsets[0]]
        prev_inhale_offset = self.local_corrected_volume[self.inhale_offsets[0]]
        prev_exhale_onset = self.local_corrected_volume[self.exhale_offsets[0]]
        prev_exhale_offset = self.local_corrected_volume[self.exhale_offsets[0]]

        for idx in range(1, self.num_breaths):
            current_inhale_onset = self.local_corrected_volume[self.inhale_onsets[idx]]
            current_inhale_offset = self.local_corrected_volume[self.inhale_offsets[idx]]
            current_exhale_onset = self.local_corrected_volume[self.exhale_offsets[idx]]
            current_exhale_offset = self.local_corrected_volume[self.exhale_offsets[idx]]

            if ((abs(current_inhale_onset - prev_inhale_onset) > threshold) or
                    (abs(current_inhale_offset - prev_inhale_offset) > threshold) or
                    (abs(current_exhale_onset - prev_exhale_onset) > threshold) or
                    (abs(current_exhale_offset - prev_exhale_offset) > threshold)):
                self.movement[idx] = True

            threshold = movement_threshold * np.mean([abs(self.inhale_volumes[idx]), abs(self.exhale_volumes[idx])])
            prev_inhale_onset = self.local_corrected_volume[self.inhale_onsets[idx]]
            prev_inhale_offset = self.local_corrected_volume[self.inhale_offsets[idx]]
            prev_exhale_onset = self.local_corrected_volume[self.exhale_offsets[idx]]
            prev_exhale_offset = self.local_corrected_volume[self.exhale_offsets[idx]]

    def find_movement_breaths(self, movement_threshold, remove_movement=True):
        self.movement = np.full(self.num_breaths, False)

        if not remove_movement:
            return

        for idx in range(self.num_breaths):
            inhale_center = np.average([self.local_corrected_volume[self.inhale_onsets[idx]],
                                        self.local_corrected_volume[self.inhale_offsets[idx]]])
            exhale_center = np.average([self.local_corrected_volume[self.exhale_onsets[idx]],
                                        self.local_corrected_volume[self.exhale_offsets[idx]]])
            if ((abs(inhale_center - self.data_zero) > movement_threshold) or
                    (abs(exhale_center - self.data_zero) > movement_threshold)):
                self.movement[idx] = True

    def find_movement_breaths_old(self, movement_threshold, remove_movement=True):
        self.movement = np.full(self.num_breaths, False)

        if not remove_movement:
            return

        min_threshold = self.data_zero - movement_threshold
        max_threshold = self.data_zero + movement_threshold
        for idx in range(self.num_breaths):
            breath_center = np.average([self.local_corrected_volume[self.inhale_onsets[idx]],
                                        self.local_corrected_volume[self.inhale_offsets[idx]],
                                        self.local_corrected_volume[self.exhale_onsets[idx]],
                                        self.local_corrected_volume[self.exhale_offsets[idx]]])
            if not min_threshold < breath_center < max_threshold:
                self.movement[idx] = True

    def find_outlier_breaths(self, outlier_test_stats, outlier_window_size, max_std, remove_outlier=False):
        self.outlier = np.full(self.num_breaths, False)
        self.outlier_piv = np.full(self.num_breaths, False)
        self.outlier_pev = np.full(self.num_breaths, False)
        self.outlier_pif = np.full(self.num_breaths, False)
        self.outlier_pef = np.full(self.num_breaths, False)
        self.outlier_ti = np.full(self.num_breaths, False)
        self.outlier_te = np.full(self.num_breaths, False)
        self.outlier_tb = np.full(self.num_breaths, False)

        if not remove_outlier:
            return

        for idx in range(self.num_breaths):
            # Determine which breaths lie within outlier detection window for each breath
            current_breath_time = self.time[self.start_idx[idx]]

            left_idx = idx
            while left_idx >= 0:
                left_time = self.time[self.start_idx[left_idx]]
                if (current_breath_time - left_time) <= (outlier_window_size / 2.0):
                    # Breath lies within outlier determination window, continue searching left
                    left_idx = left_idx - 1
                else:
                    # Stop searching
                    left_idx = left_idx + 1
                    break
            if left_idx < 0:
                left_idx = 0

            right_idx = idx
            while right_idx < self.num_breaths:
                right_time = self.time[self.start_idx[right_idx]]
                if (right_time - current_breath_time) <= (outlier_window_size / 2.0):
                    # Breath lies within outlier determination window, continue searching right
                    right_idx = right_idx + 1
                else:
                    # Stop searching
                    break

            troubleshooting = False
            if troubleshooting:
                print('left_idx: %s, left_time: %s' % (left_idx, self.time[self.start_idx[left_idx]]))
                print('current_idx: %s, current_time: %s' % (idx, current_breath_time))
                print('right_idx: %s, right_time: %s' % (right_idx, self.time[self.start_idx[right_idx - 1]]))
                if left_idx > 0:
                    print('next_left_time:', self.time[self.start_idx[left_idx - 1]])
                if right_idx < self.num_breaths:
                    print('next_right_time:', self.time[self.start_idx[right_idx]])
                print()

            # Check breath against each outlier stat to determine if it is an outlier
            if 'piv' in outlier_test_stats:
                # Find mean and standard deviation of nearby breaths
                avg_piv = np.mean(self.inhale_volumes[left_idx:right_idx])
                std_piv = np.std(self.inhale_volumes[left_idx:right_idx])

                # Determine if current breath is an outlier
                diff = abs(self.inhale_volumes[idx] - avg_piv)
                if diff > max_std * std_piv:
                    self.outlier_piv[idx] = True
                    self.outlier[idx] = True

            if 'pev' in outlier_test_stats:
                # Find mean and standard deviation of nearby breaths
                avg_pev = np.mean(self.exhale_volumes[left_idx:right_idx])
                std_pev = np.std(self.exhale_volumes[left_idx:right_idx])

                # Determine if current breath is an outlier
                diff = abs(self.exhale_volumes[idx] - avg_pev)
                if diff > max_std * std_pev:
                    self.outlier_pev[idx] = True
                    self.outlier[idx] = True

            if 'pif' in outlier_test_stats:
                # Find mean and standard deviation of nearby breaths
                avg_pif = np.mean(self.inhale_peak_flows[left_idx:right_idx])
                std_pif = np.std(self.inhale_peak_flows[left_idx:right_idx])

                # Determine if current breath is an outlier
                diff = abs(self.inhale_peak_flows[idx] - avg_pif)
                if diff > max_std * std_pif:
                    self.outlier_pif[idx] = True
                    self.outlier[idx] = True

            if 'pef' in outlier_test_stats:
                # Find mean and standard deviation of nearby breaths
                avg_pef = np.mean(self.exhale_peak_flows[left_idx:right_idx])
                std_pef = np.std(self.exhale_peak_flows[left_idx:right_idx])

                # Determine if current breath is an outlier
                diff = abs(self.exhale_peak_flows[idx] - avg_pef)
                if diff > max_std * std_pef:
                    self.outlier_pef[idx] = True
                    self.outlier[idx] = True

            if 'ti' in outlier_test_stats:
                # Find mean and standard deviation of nearby breaths
                avg_ti = np.mean(self.inhale_times[left_idx:right_idx])
                std_ti = np.std(self.inhale_times[left_idx:right_idx])

                # Determine if current breath is an outlier
                diff = abs(self.inhale_times[idx] - avg_ti)
                if diff > max_std * std_ti:
                    self.outlier_ti[idx] = True
                    self.outlier[idx] = True

            if 'te' in outlier_test_stats:
                # Find mean and standard deviation of nearby breaths
                avg_te = np.mean(self.exhale_times[left_idx:right_idx])
                std_te = np.std(self.exhale_times[left_idx:right_idx])

                # Determine if current breath is an outlier
                diff = abs(self.exhale_times[idx] - avg_te)
                if diff > max_std * std_te:
                    self.outlier_te[idx] = True
                    self.outlier[idx] = True

            if 'tb' in outlier_test_stats:
                # Find mean and standard deviation of nearby breaths
                avg_tb = np.mean(self.breath_times[left_idx:right_idx])
                std_tb = np.std(self.breath_times[left_idx:right_idx])

                # Determine if current breath is an outlier
                diff = abs(self.breath_times[idx] - avg_tb)
                if diff > max_std * std_tb:
                    self.outlier_tb[idx] = True
                    self.outlier[idx] = True

    def find_relaxation_idx(self, data, start_idx, stop_idx, percentage):
        start_volume = self.local_corrected_volume[start_idx]
        stop_volume = self.local_corrected_volume[stop_idx]
        target_volume = (1.0 - float(percentage) / 100) * start_volume + (float(percentage) / 100) * stop_volume
        if start_volume > stop_volume:
            # Exhale
            relaxation_idx = np.argmax(data[start_idx:stop_idx + 1] <= target_volume)
        else:
            # Inhale
            relaxation_idx = np.argmax(data[start_idx:stop_idx + 1] >= target_volume)

        # adjust index to account for position in total data set
        relaxation_idx = relaxation_idx + start_idx

        return relaxation_idx

    def find_saturated_breaths(self, min_voltage, max_voltage, remove_saturated=False):
        self.saturated = np.full(self.num_breaths, False)

        if not remove_saturated:
            return

        for idx in range(self.num_breaths):
            if ((self.voltage[self.inhale_volume_peaks[idx]] >= max_voltage) or
                    (self.voltage[self.exhale_volume_peaks[idx]] <= min_voltage)):
                self.saturated[idx] = True

    def load_data_file(self, filename, max_data_points=0):
        if filename:
            # Choose which file reading method to use
            extension = filename[filename.rfind('.')::]
            if extension == '.rcalv':
                self.read_labview_datafile(filename, max_data_points)
            elif extension == '.csv':
                self.read_daqami_datafile(filename, max_data_points)
            else:
                print('File type not recognized')
        return

    def lowpass_filter_pressure(self, cutoff, order):
        # Filter pressure signal with a lowpass Butterworth filter
        filtered = butter_filter(np.asarray(self.pressure), cutoff, self.sampling_rate, 'lowpass', order)
        self.lowpass_pressure = filtered

    @staticmethod
    def parse_header_line(header_line):
        # Given a header line in the form: ['Parameter: value'], returns the value of the parameter as a String object
        return header_line.split(": ")[1]

    def read_calibration_datafile(self, filename):
        wb = openpyxl.load_workbook(filename=filename, read_only=True)
        ws = wb['Calibration Data']

        m = float()
        b = float()
        r_squared = float()
        times = []
        volumes = []
        pressures = []
        for row in ws.rows:
            if row[0].value == 'Calibration Slope [Pa/mL]':
                m = row[1].value
            elif row[0].value == 'Calibration Intercept [Pa]':
                b = row[1].value
            elif row[0].value == 'Calibration R^2':
                r_squared = row[1].value
            elif row[0].value == 'Time [s]':
                pass
            else:
                t = row[0].value
                v = row[1].value
                p = row[2].value
                times.append(t)
                volumes.append(v)
                pressures.append(p)

        if not r_squared:
            # Calculate r_squared value
            y_mean = np.average(pressures)
            unexplained_error = 0
            total_error = 0
            for idx in range(len(volumes)):
                y = pressures[idx]
                y_pred = m * volumes[idx] + b
                unexplained_error += (y - y_pred) ** 2
                total_error += (y - y_mean) ** 2
            r_squared = 1 - (unexplained_error / total_error)

        self.calibration_m = m
        self.calibration_b = b
        self.calibration_r_squared = r_squared
        self.calibration_times = tuple(list(times))
        self.calibration_volumes = tuple(list(volumes))
        self.calibration_pressures = tuple(list(pressures))
        min_pressure = wb['Analysis Settings']['B1'].value
        max_pressure = wb['Analysis Settings']['B2'].value
        ref_pressure = wb['Analysis Settings']['B3'].value

        return min_pressure, max_pressure, ref_pressure

    # Todo: update how read methods respond in case where empty lines exist at end of datafile (i.e. file saved badly)
    def read_labview_datafile(self, filename, max_data_points=0):
        # Reads labview .rcalv file containing mouse respiration data into Respiration object
        with open(filename) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=",")
            channel_count = 0
            sample_count = 0
            start_time = 0
            sampling_rate = 0
            data_idx = []
            time = []
            voltage = []
            comments = []
            comment_idxs = []
            line_count = 0
            for row in csv_reader:
                if (line_count > (max_data_points + 6)) and (max_data_points != 0):
                    print("Max data limit reached")
                    break
                elif line_count == 2:
                    # Channel Count
                    channel_count = int(self.parse_header_line(row[0]))
                elif line_count == 4:
                    start_time = self.parse_header_line(row[0])
                elif line_count == 5:
                    # Sampling rate
                    sampling_rate = float(self.parse_header_line(row[0]))
                elif line_count >= 7:
                    # Check for datafile irregularities
                    if len(row) != 3:
                        print("\nWarning - Datafile irregularity found:")
                        print("line_count:", line_count, "row:", row)
                    else:
                        # Data values
                        data_idx.append(sample_count)
                        time.append(float(row[0]))
                        voltage.append(float(row[1]))
                        comments.append(row[2])
                        if row[2]:
                            comment_idxs.append(sample_count)
                        sample_count += 1
                line_count += 1

        self.channel_count = channel_count
        self.sample_count = sample_count
        self.start_time = start_time
        self.sampling_rate = sampling_rate
        self.data_idx = np.array(data_idx)
        self.time = np.array(time)
        self.voltage = np.array(voltage)
        self.comments = np.array(comments)
        self.comment_idxs = np.array(comment_idxs)
        self.len_data = len(time)

    def read_daqami_datafile(self, filename, max_data_points=0):
        # Reads daqami .csv file containing mouse respiration data into Respiration object
        with open(filename) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=",")
            channel_count = 0
            sample_count = 0
            start_time = 0
            sampling_rate = 0
            data_idx = []
            time = []
            voltage = []
            line_count = 0
            for row in csv_reader:
                if (line_count > (max_data_points + 6)) and (max_data_points != 0):
                    print("Max data limit reached")
                    break
                elif line_count == 2:
                    # Channel Count
                    channel_count = int(self.parse_header_line(row[0]))
                elif line_count == 3:
                    # Sample Count
                    sample_count = int(self.parse_header_line(row[0]))
                elif line_count == 4:
                    start_time = self.parse_header_line(row[0])
                elif line_count == 5:
                    # Sampling rate
                    sampling_rate = float(self.parse_header_line(row[0]))
                elif line_count >= 7:
                    # Data values
                    data_idx.append(int(row[0]))
                    time.append(round(int(row[0]) * (1 / sampling_rate), 3))
                    voltage.append(float(row[2]))
                line_count += 1

        self.channel_count = channel_count
        self.sample_count = sample_count
        self.start_time = start_time
        self.sampling_rate = sampling_rate
        self.data_idx = np.array(data_idx)
        self.time = np.array(time)
        self.voltage = np.array(voltage)
        self.len_data = len(time)

    def remove_calibration_point(self, idx):
        self.calibration_times = np.delete(self.calibration_times, idx)
        self.calibration_pressures = np.delete(self.calibration_pressures, idx)
        self.calibration_volumes = np.delete(self.calibration_volumes, idx)

    def remove_global_drift(self, time, data):
        # Remove global drift using linear regression on whole data set
        [m, b] = np.polyfit(time, data, 1)
        self.global_corrected_volume = data - (m * time + b)

    def remove_local_drift(self, data, window_size):
        window_size = round(window_size * self.sampling_rate)  # convert window size from seconds to # of data points
        # mean kernel of size = window_size, normalized to 1
        kernel = [1.0 / window_size] * window_size

        correction = np.convolve(data, kernel, 'same')
        corrected_data = data - correction
        self.local_corrected_volume = corrected_data
        return correction, corrected_data

    def smooth_volume_signal(self, sigma=2.0):
        # Smooth volume signal with Gaussian filter, default sigma=2.0
        self.smoothed_volume = smoothing(self.volume, sigma)
        return self.smoothed_volume

    def write_data_to_xlsx(self, filename, analysis_window):
        wb = openpyxl.Workbook(write_only=True)

        ws_stat_avgs = wb.create_sheet('1 min Stat Averages', 0)
        ws_stat_avgs_headings = ['Window Number',
                                 'Number of Breaths Detected',
                                 'Number of Breaths Used',
                                 'Number of Breaths Excluded',
                                 'Number of Saturated Breaths',
                                 'Number of Movement Breaths',
                                 'Number of Outlier Breaths',
                                 'Absolute Window Start Time [s]',
                                 'Absolute Window Stop Time [s]',
                                 'Relative Window Center Time [s]',
                                 'Relative Window Center Time [min]',
                                 'Average Inspiratory Volume [mL]', '% of Baseline',
                                 'Average Expiratory Volume [mL]', '% of Baseline',
                                 'Average Duration of Inspiration [s]', '% of Baseline',
                                 'Average Duration of Expiration [s]', '% of Baseline',
                                 'Average Breaths per Minute [breaths/min]', '% of Baseline',
                                 'Average Enhanced Pause (Penh)', '% of Baseline',
                                 'Average Mid-Expiratory Flow [mL/s]', '% of Baseline',
                                 'Average Duration of Post Inspiratory Pause [s]', '% of Baseline',
                                 'Average Duration of PIP + Expiration [s]', '% of Baseline',
                                 'Average PIP + Expiration Volume [s]', '% of Baseline',
                                 'Average Post Inspiratory Pause Volume [mL]', '% of Baseline',
                                 'Average Post Expiratory Pause Volume [mL]', '% of Baseline',
                                 'Average Peak Inspiratory Flow [mL/s]', '% of Baseline',
                                 'Average Peak Expiratory Flow [mL/s]', '% of Baseline',
                                 'Average Mid-Inspiratory Flow [mL/s]', '% of Baseline',
                                 'Average Duration of Post Expiratory Pause [s]', '% of Baseline',
                                 'Average Total Duration of Breath [s]', '% of Baseline',
                                 'Average Relaxation Time to 50% Inspiratory Volume [s]', '% of Baseline',
                                 'Average Relaxation Time to 64% Inspiratory Volume [s]', '% of Baseline',
                                 'Average Relaxation Time to 50% Expiratory Volume [s]', '% of Baseline',
                                 'Average Relaxation Time to 64% Expiratory Volume [s]', '% of Baseline']

        ws_stat_avgs.append(ws_stat_avgs_headings)

        # Convert window_rel_center_times to minutes
        window_rel_center_time_minutes = self.window_rel_center_time / 60.0

        for idx in range(self.num_windows):
            row = [idx,
                   self.window_num_breaths[idx],
                   self.window_num_breaths[idx] - self.window_num_excluded[idx],
                   self.window_num_excluded[idx],
                   self.window_num_saturated[idx],
                   self.window_num_movement[idx],
                   self.window_num_outlier[idx],
                   self.window_start_time[idx],
                   self.window_stop_time[idx],
                   self.window_rel_center_time[idx],
                   window_rel_center_time_minutes[idx],  # Report time in minutes
                   self.avg_inhale_volume[idx], '',
                   abs(self.avg_exhale_volume[idx]), '',  # Report as positive number
                   self.avg_inhale_time[idx], '',
                   self.avg_exhale_time[idx], '',
                   self.avg_bpm[idx], '',
                   self.avg_penh[idx], '',
                   abs(self.avg_ef50[idx]), '',  # Report as positive number
                   self.avg_inhale_pause_time[idx], '',
                   self.avg_inhale_pause_plus_exhale_time[idx], '',
                   self.avg_inhale_pause_plus_exhale_volume[idx], '',
                   self.avg_inhale_pause_volume[idx], '',
                   self.avg_exhale_pause_volume[idx], '',
                   self.avg_inhale_peak_flow[idx], '',
                   abs(self.avg_exhale_peak_flow[idx]), '',  # Report as positive number
                   self.avg_if50[idx], '',
                   self.avg_exhale_pause_time[idx], '',
                   self.avg_breath_time[idx], '',
                   self.avg_tri_50[idx], '',
                   self.avg_tri_64[idx], '',
                   self.avg_tre_50[idx], '',
                   self.avg_tre_64[idx], '']

            ws_stat_avgs.append(row)

        # Find min/max for each stat average across all windows
        min_avg_inhale_volume = np.argmin(self.avg_inhale_volume)
        min_avg_exhale_volume = np.argmin(abs(self.avg_exhale_volume))
        max_avg_inhale_time = np.argmax(self.avg_inhale_time)
        max_avg_exhale_time = np.argmax(self.avg_exhale_time)
        min_avg_bpm = np.argmin(self.avg_bpm)
        max_avg_penh = np.argmax(self.avg_penh)
        min_avg_ef50 = np.argmin(abs(self.avg_ef50))
        max_avg_inhale_pause_time = np.argmax(self.avg_inhale_pause_time)
        max_avg_inhale_pause_plus_exhale_time = np.argmax(self.avg_inhale_pause_plus_exhale_time)
        min_avg_inhale_pause_plus_exhale_volume = np.argmin(abs(self.avg_inhale_pause_plus_exhale_volume))

        min_max_value_row = ['', '', '', '', '', '', '', '', '', '', 'Min/Max',
                             self.avg_inhale_volume[min_avg_inhale_volume], '',
                             abs(self.avg_exhale_volume)[min_avg_exhale_volume], '',
                             self.avg_inhale_time[max_avg_inhale_time], '',
                             self.avg_exhale_time[max_avg_exhale_time], '',
                             self.avg_bpm[min_avg_bpm], '',
                             self.avg_penh[max_avg_penh], '',
                             abs(self.avg_ef50)[min_avg_ef50], '',
                             self.avg_inhale_pause_time[max_avg_inhale_pause_time], '',
                             self.avg_inhale_pause_plus_exhale_time[max_avg_inhale_pause_plus_exhale_time], '',
                             self.avg_inhale_pause_plus_exhale_volume[min_avg_inhale_pause_plus_exhale_volume], '']

        min_max_time_row = ['', '', '', '', '', '', '', '', '', '', 'Time',
                            window_rel_center_time_minutes[min_avg_inhale_volume], '',
                            window_rel_center_time_minutes[min_avg_exhale_volume], '',
                            window_rel_center_time_minutes[max_avg_inhale_time], '',
                            window_rel_center_time_minutes[max_avg_exhale_time], '',
                            window_rel_center_time_minutes[min_avg_bpm], '',
                            window_rel_center_time_minutes[max_avg_penh], '',
                            window_rel_center_time_minutes[min_avg_ef50], '',
                            window_rel_center_time_minutes[max_avg_inhale_pause_time], '',
                            window_rel_center_time_minutes[max_avg_inhale_pause_plus_exhale_time], '',
                            window_rel_center_time_minutes[min_avg_inhale_pause_plus_exhale_volume], '']

        ws_stat_avgs.append([])
        ws_stat_avgs.append(min_max_value_row)
        ws_stat_avgs.append(min_max_time_row)

        ws_stat_stdevs = wb.create_sheet('1 min Stat Std. Dev.', 1)
        ws_stat_stdevs_headings = ['Window Number',
                                   'Number of Breaths Detected',
                                   'Number of Breaths Used',
                                   'Number of Breaths Excluded',
                                   'Number of Saturated Breaths',
                                   'Number of Movement Breaths',
                                   'Number of Outlier Breaths',
                                   'Absolute Window Start Time [s]',
                                   'Absolute Window Stop Time [s]',
                                   'Relative Window Center Time [s]',
                                   'Relative Window Center Time [min]',
                                   'Std. Dev. Inspiratory Volume [mL]',
                                   'Std. Dev. Expiratory Volume [mL]',
                                   'Std. Dev. Duration of Inspiration [s]',
                                   'Std. Dev. Duration of Expiration [s]',
                                   'Std. Dev. Breaths per Minute [breaths/min]',
                                   'Std. Dev. Enhanced Pause (Penh)',
                                   'Std. Dev. Mid-Expiratory Flow [mL/s]',
                                   'Std. Dev. Duration of Post Inspiratory Pause [s]',
                                   'Std. Dev. Duration of PIP + Expiration [s]',
                                   'Std. Dev. PIP + Expiration Volume [s]',
                                   'Std. Dev. Post Inspiratory Pause Volume [mL]',
                                   'Std. Dev. Post Expiratory Pause Volume [mL]',
                                   'Std. Dev. Peak Inspiratory Flow [mL/s]',
                                   'Std. Dev. Peak Expiratory Flow [mL/s]',
                                   'Std. Dev. Mid-Inspiratory Flow [mL/s]',
                                   'Std. Dev. Duration of Post Expiratory Pause [s]',
                                   'Std. Dev. Total Duration of Breath [s]',
                                   'Std. Dev. Relaxation Time to 50% Inspiratory Volume [s]',
                                   'Std. Dev. Relaxation Time to 64% Inspiratory Volume [s]',
                                   'Std. Dev. Relaxation Time to 50% Expiratory Volume [s]',
                                   'Std. Dev. Relaxation Time to 64% Expiratory Volume [s]']

        ws_stat_stdevs.append(ws_stat_stdevs_headings)

        for idx in range(self.num_windows):
            row = [idx,
                   self.window_num_breaths[idx],
                   self.window_num_breaths[idx] - self.window_num_excluded[idx],
                   self.window_num_excluded[idx],
                   self.window_num_saturated[idx],
                   self.window_num_movement[idx],
                   self.window_num_outlier[idx],
                   self.window_start_time[idx],
                   self.window_stop_time[idx],
                   self.window_rel_center_time[idx],
                   window_rel_center_time_minutes[idx],  # Report time in minutes
                   self.stdev_inhale_volume[idx],
                   self.stdev_exhale_volume[idx],  # Report as positive number
                   self.stdev_inhale_time[idx],
                   self.stdev_exhale_time[idx],
                   self.stdev_bpm[idx],
                   self.stdev_penh[idx],
                   self.stdev_ef50[idx],  # Report as positive number
                   self.stdev_inhale_pause_time[idx],
                   self.stdev_inhale_pause_plus_exhale_time[idx],
                   self.stdev_inhale_pause_plus_exhale_volume[idx],
                   self.stdev_inhale_pause_volume[idx],
                   self.stdev_exhale_pause_volume[idx],
                   self.stdev_inhale_peak_flow[idx],
                   self.stdev_exhale_peak_flow[idx],  # Report as positive number
                   self.stdev_if50[idx],
                   self.stdev_exhale_pause_time[idx],
                   self.stdev_breath_time[idx],
                   self.stdev_tri_50[idx],
                   self.stdev_tri_64[idx],
                   self.stdev_tre_50[idx],
                   self.stdev_tre_64[idx]]

            ws_stat_stdevs.append(row)

        stat_avgs = [np.mean(self.inhale_volumes),
                     abs(np.mean(self.exhale_volumes)),  # Report as positive number
                     np.mean(self.inhale_times),
                     np.mean(self.exhale_times),
                     np.mean(self.bpm),
                     np.mean(self.penh),
                     abs(np.mean(self.ef50)),  # Report as positive number
                     np.mean(self.inhale_pause_times),
                     np.mean(self.inhale_pause_plus_exhale_times),
                     np.mean(self.inhale_pause_plus_exhale_volumes),
                     np.mean(self.inhale_pause_volumes),
                     np.mean(self.exhale_pause_volumes),
                     np.mean(self.inhale_peak_flows),
                     abs(np.mean(self.exhale_peak_flows)),  # Report as positive number
                     np.mean(self.if50),
                     np.mean(self.exhale_pause_times),
                     np.mean(self.breath_times),
                     np.mean(self.tri_50),
                     np.mean(self.tri_64),
                     np.mean(self.tre_50),
                     np.mean(self.tre_64)]

        ws_overall_stat_avgs = wb.create_sheet('Overall Stat Averages', 2)
        ws_overall_stat_avgs_headings = ['Measurement', 'Overall Averages']
        ws_overall_stat_avgs_stat_titles = ['Inspiratory Volume [mL]',
                                            'Expiratory Volume [mL]',
                                            'Duration of Inspiration [s]',
                                            'Duration of Expiration [s]',
                                            'Breaths per Minute [breaths/min]',
                                            'Enhanced Pause (Penh)',
                                            'Mid-Expiratory Flow [mL/s]',
                                            'Duration of Post Inspiratory Pause [s]',
                                            'Duration of PIP + Expiration [s]',
                                            'PIP + Expiration Volume [mL]',
                                            'Post Inspiratory Pause Volume [mL]',
                                            'Post Expiratory Pause Volume [mL]',
                                            'Peak Inspiratory Flow [mL/s]',
                                            'Peak Expiratory Flow [mL/s]',
                                            'Mid-Inspiratory Flow [mL/s]',
                                            'Duration of Post Expiratory Pause [s]',
                                            'Total Breath Duration [s]',
                                            'Relaxation Time to 50% Inspiratory Volume [s]',
                                            'Relaxation Time to 64% Inspiratory Volume [s]',
                                            'Relaxation Time to 50% Expiratory Volume [s]',
                                            'Relaxation Time to 64% Expiratory Volume [s]']

        ws_overall_stat_avgs.append(ws_overall_stat_avgs_headings)
        for idx in range(len(stat_avgs)):
            # Report peak expiratory volume and mid-expiratory flow as positive values
            row = [ws_overall_stat_avgs_stat_titles[idx], stat_avgs[idx]]
            ws_overall_stat_avgs.append(row)

        ws_stats = wb.create_sheet('Breath Statistics', 3)
        ws_stats_headings = ['Breath Number',
                             'Excluded?',
                             'Saturated?',
                             'Movement?',
                             'Outlier?',
                             'Inspiratory Volume Outlier?',
                             'Expiratory Volume Outlier?',
                             'Peak Inspiratory Flow Outlier?',
                             'Peak Expiratory Flow Outlier?',
                             'Duration of Inspiration Outlier?',
                             'Duration of Expiration Outlier?',
                             'Total Breath Duration Outlier?',
                             'Breath Onset [s]',
                             'Breath Offset [s]',
                             'Inspiratory Volume [mL]',
                             'Post Inspiratory Pause Volume [mL]',
                             'Expiratory Volume [mL]',
                             'Post Expiratory Pause Volume [mL]',
                             'PIP + Expiratory Volume [mL]',
                             'Peak Inspiratory Flow [mL/s]',
                             'Peak Expiratory Flow [mL/s]',
                             'Mid-Inspiratory Flow [mL/s]',
                             'Mid-Expiratory Flow [mL/s]',
                             'Duration of Inspiration [s]',
                             'Duration of Post Inspiratory Pause [s]',
                             'Duration of Expiration [s]',
                             'Duration of Post Expiratory Pause [s]',
                             'Duration of PIP + Expiration [s]',
                             'Total Breath Duration [s]',
                             'Breaths per Minute [breaths/min]',
                             'Relaxation Time to 50% Inspiratory Volume [s]',
                             'Relaxation Time to 64% Inspiratory Volume [s]',
                             'Relaxation Time to 50% Expiratory Volume [s]',
                             'Relaxation Time to 64% Expiratory Volume [s]',
                             'Enhanced Pause (Penh)']

        ws_stats.append(ws_stats_headings)
        for idx in range(self.num_breaths):
            row = [idx,
                   self.breaths_to_exclude[idx],
                   self.saturated[idx],
                   self.movement[idx],
                   self.outlier[idx],
                   self.outlier_piv[idx],
                   self.outlier_pev[idx],
                   self.outlier_pif[idx],
                   self.outlier_pef[idx],
                   self.outlier_ti[idx],
                   self.outlier_te[idx],
                   self.outlier_tb[idx],
                   self.time[self.start_idx.data[idx]],
                   self.time[self.stop_idx.data[idx]],
                   self.inhale_volumes.data[idx],
                   self.inhale_pause_volumes.data[idx],
                   self.exhale_volumes.data[idx],
                   self.exhale_pause_volumes.data[idx],
                   self.inhale_pause_plus_exhale_volumes.data[idx],
                   self.inhale_peak_flows.data[idx],
                   self.exhale_peak_flows.data[idx],
                   self.if50.data[idx],
                   self.ef50.data[idx],
                   self.inhale_times.data[idx],
                   self.inhale_pause_times.data[idx],
                   self.exhale_times.data[idx],
                   self.exhale_pause_times.data[idx],
                   self.inhale_pause_plus_exhale_times.data[idx],
                   self.breath_times.data[idx],
                   self.bpm.data[idx],
                   self.tri_50.data[idx],
                   self.tri_64.data[idx],
                   self.tre_50.data[idx],
                   self.tre_64.data[idx],
                   self.penh.data[idx]]

            ws_stats.append(row)

        ws_landmarks = wb.create_sheet('Breath Landmarks', 4)
        ws_landmarks_headings = ['Breath Number',
                                 'Excluded?',
                                 'Breath Onset',
                                 'Breath Offset',
                                 'Inhale Onset',
                                 'Inhale Offset',
                                 'Inhale Pause Onset',
                                 'Inhale Pause Offset',
                                 'Exhale Onset',
                                 'Exhale Offset',
                                 'Exhale Pause Onset',
                                 'Exhale Pause Offset',
                                 'Inhale Flow Peak',
                                 'Exhale Flow Peak',
                                 'Inhale Volume Peak',
                                 'Exhale Volume Peak',
                                 '50% Inhale Volume',
                                 '64% Inhale Volume',
                                 '50% Exhale Volume',
                                 '64% Exhale Volume']

        ws_landmarks.append(ws_landmarks_headings)
        for idx in range(self.num_breaths):
            row = [idx,
                   self.breaths_to_exclude[idx],
                   self.start_idx.data[idx],
                   self.stop_idx.data[idx],
                   self.inhale_onsets.data[idx],
                   self.inhale_offsets.data[idx],
                   self.inhale_pause_onsets.data[idx],
                   self.inhale_pause_offsets.data[idx],
                   self.exhale_onsets.data[idx],
                   self.exhale_offsets.data[idx],
                   self.exhale_pause_onsets.data[idx],
                   self.exhale_pause_offsets.data[idx],
                   self.inhale_flow_peaks.data[idx],
                   self.exhale_flow_peaks.data[idx],
                   self.inhale_volume_peaks.data[idx],
                   self.exhale_volume_peaks.data[idx],
                   self.i50v.data[idx],
                   self.i64v.data[idx],
                   self.e50v.data[idx],
                   self.e64v.data[idx]]

            ws_landmarks.append(row)

        ws_calibration = wb.create_sheet('Calibration Data', 5)
        ws_calibration.append(['Calibration Slope [Pa/mL]', self.calibration_m])
        ws_calibration.append(['Calibration Intercept [Pa]', self.calibration_b])
        ws_calibration.append(['Calibration R^2', self.calibration_r_squared])
        ws_calibration_headings = ['Time [s]', 'Injection Volume [mL]', 'Pressure [Pa]']
        ws_calibration.append(ws_calibration_headings)
        for idx in range(len(self.calibration_times)):
            row = [self.calibration_times[idx], self.calibration_volumes[idx], self.calibration_pressures[idx]]
            ws_calibration.append(row)

        ws_settings = wb.create_sheet('Analysis Settings', 6)
        ws_settings_headings = ['DAQ Minimum Pressure [Pa]',
                                'DAQ Maximum Pressure [Pa]',
                                'DAQ Reference Pressure [Pa]',
                                'DAQ Supply Voltage [V]',
                                'DAQ Channel Count',
                                'DAQ Sample Count',
                                'DAQ Start Time [s]',
                                'DAQ Sampling Rate [Hz]',
                                'Software Version',
                                'Use Previous Calibration Data',
                                'Previous Calibration Data Filename',
                                'Lower Saturation Voltage [V]',
                                'Upper Saturation Voltage [V]',
                                'Lowpass Filter Cutoff Frequency [Hz]',
                                'Lowpass Filter Order',
                                'Autocalibration Gaussian Filter Smoothing Sigma [s]',
                                'Autocalibration Seconds to Search Left of Comment [s]',
                                'Autocalibration Seconds to Search Right of Comment [s]',
                                'Autocalibration Duplicate Comment Distance [s]',
                                'Autocalibration Lowest Calibration Volume [mL]',
                                'Autocalibration Highest Calibration Volume [mL]',
                                'Autocalibration Reference Calibration Volume [mL]',
                                'Autocalibration Step Window Width for Reference Calibration Volume [s]',
                                'Autocalibration Variance Window Width [s]',
                                'Autocalibration Number of Standard Deviations for Step Height Threshold',
                                'Autocalibration Step Width Threshold [s]',
                                'Autocalibration Lower Variance Ratio',
                                'Autocalibration Downsampling Rate',
                                'Autocalibration Number of Standard Deviations for Calibration Point Exclusion',
                                'Volume Signal Gaussian Smoothing Filter Sigma',
                                'Local Drift Correction Window Size [s]',
                                'Flow Signal Gaussian Smoothing Filter Sigma',
                                'Peak Detection Window Sizes [samples]',
                                'Peak Detection Shift Sizes',
                                'Pause Search Volume Threshold',
                                'Pause Detection Number of Histograms',
                                'Pause Detection Extra Zero Bin Threshold',
                                'Pause Detection Extra Pause Bin Threshold',
                                'Pause Detection Mode Bin Ratio',
                                'Pause Detection Extra Pause Bin Ratio',
                                'Movement Threshold [mL]',
                                'Breath Exclusion Adjacency',
                                'Breath Exclusion Must Be Between Exclusions',
                                'Region of Interest Start [s]',
                                'Region of Interest Stop [s]',
                                'Size of Averaging Window [s]',
                                'Offset Between Averaging Windows [s]',
                                'Check for Peak Inspiratory Volume Outliers',
                                'Check for Peak Expiratory Volume Outliers',
                                'Check for Peak Inspiratory Flow Outliers',
                                'Check for Peak Expiratory Flow Outliers',
                                'Check for Duration of Inspiration Outliers',
                                'Check for Duration of Expiration Outliers',
                                'Check for Duration of Breath Outliers',
                                'Outlier Determination Window Size [s]',
                                'Outlier Maximum Number of Standard Deviations from Mean',
                                'Save Signal Data',
                                'Remove Breaths with Voltage Saturation',
                                'Remove Breaths where Movement Detected',
                                'Remove Outlier Breaths']

        ws_settings_row = [analysis_window.min_pressure,
                           analysis_window.max_pressure,
                           analysis_window.ref_pressure,
                           analysis_window.supply_voltage,
                           self.channel_count,
                           self.sample_count,
                           self.start_time,
                           self.sampling_rate,
                           analysis_window.version,
                           analysis_window.use_prev_cal,
                           analysis_window.prev_cal_filename,
                           analysis_window.sat_min_voltage,
                           analysis_window.sat_max_voltage,
                           analysis_window.volume_calibration_cutoff,
                           analysis_window.volume_calibration_order,
                           analysis_window.autocal_sigma,
                           analysis_window.autocal_search_left,
                           analysis_window.autocal_search_right,
                           analysis_window.autocal_dup_comment_distance,
                           analysis_window.autocal_lowest_cal_vol,
                           analysis_window.autocal_highest_cal_vol,
                           analysis_window.autocal_ref_cal_vol,
                           analysis_window.autocal_w_ref_cal_vol,
                           analysis_window.autocal_w_variance,
                           analysis_window.autocal_num_std_step_height_threshold,
                           analysis_window.autocal_step_width_threshold,
                           analysis_window.autocal_lower_variance_ratio,
                           analysis_window.autocal_downsampling_rate,
                           analysis_window.autocal_num_std_exclude,
                           analysis_window.volume_smoothing_sigma,
                           analysis_window.local_drift_correction_window_size,
                           analysis_window.flow_oversmoothing_sigma,
                           str(analysis_window.peak_detection_window_sizes).strip('[]'),
                           str(analysis_window.peak_detection_shift_sizes).strip('[]'),
                           analysis_window.pause_search_threshold,
                           analysis_window.number_histogram_bins,
                           analysis_window.extra_zero_bin_threshold,
                           analysis_window.extra_pause_bin_threshold,
                           analysis_window.mode_bin_ratio,
                           analysis_window.extra_pause_bin_ratio,
                           analysis_window.movement_threshold,
                           analysis_window.exclusion_adjacency,
                           analysis_window.must_be_between_exclusions,
                           analysis_window.roi_start,
                           analysis_window.roi_stop,
                           analysis_window.avg_window_size,
                           analysis_window.avg_window_offset,
                           'piv' in analysis_window.outlier_test_stats,
                           'pev' in analysis_window.outlier_test_stats,
                           'pif' in analysis_window.outlier_test_stats,
                           'pef' in analysis_window.outlier_test_stats,
                           'ti' in analysis_window.outlier_test_stats,
                           'te' in analysis_window.outlier_test_stats,
                           'tb' in analysis_window.outlier_test_stats,
                           analysis_window.outlier_window_size,
                           analysis_window.max_num_stdevs,
                           analysis_window.save_signal_data.get(),
                           analysis_window.remove_saturated.get(),
                           analysis_window.remove_movement.get(),
                           analysis_window.remove_outlier.get()]

        for idx in range(len(ws_settings_headings)):
            row = [ws_settings_headings[idx], ws_settings_row[idx]]
            ws_settings.append(row)

        if analysis_window.save_signal_data.get():
            ws_signals = wb.create_sheet('Signal Data', 7)
            ws_signals_headings = ['Sample',
                                   'Time [s]',
                                   'Comments',
                                   'Voltage [V]',
                                   'Pressure [Pa]',
                                   'Lowpass Filtered Pressure [Pa]',
                                   'Volume [mL]',
                                   'Smoothed Volume [mL]',
                                   'Global Corrected Volume [mL]',
                                   'Local Corrected Volume [mL]',
                                   'Volumetric Flow [mL/s]']
            ws_signals.append(ws_signals_headings)
            for idx in range(analysis_window.roi_start_idx, analysis_window.roi_stop_idx):
                row = [self.data_idx[idx],
                       self.time[idx],
                       self.comments[idx],
                       self.voltage[idx],
                       self.pressure[idx],
                       self.lowpass_pressure[idx],
                       self.volume[idx],
                       self.smoothed_volume[idx],
                       self.global_corrected_volume[idx],
                       self.local_corrected_volume[idx],
                       self.flow[idx]]

                ws_signals.append(row)

        wb.save(filename=filename)

    def moving_step_fit(self, signal, sigma, search_left, search_right, dup_comment_distance, lowest_cal_vol,
                        highest_cal_vol, ref_cal_vol, w_ref_cal_volume, w_variance, num_std_step_height_threshold,
                        step_width_threshold, lower_variance_ratio, downsampling_rate, num_std_exclude,
                        make_plots=False):
        # downsampling should be integer in set: (1, 2, 3, 4, 5, 10)

        # # try removing local drift before calibration
        # window_size_sec = 2.0
        # window_size = round(window_size_sec * self.sampling_rate)
        # kernel = [1.0 / window_size] * window_size
        #
        # correction = np.convolve(signal, kernel, 'same')
        # signal = signal - correction

        # Adjust parameters based on sampling rate
        sigma = int(sigma * self.sampling_rate)
        search_left = int(search_left * self.sampling_rate)
        search_right = int(search_right * self.sampling_rate)
        dup_comment_distance = int(dup_comment_distance * self.sampling_rate)
        w_ref_cal_volume = int(w_ref_cal_volume * self.sampling_rate)
        w_variance = int(w_variance * self.sampling_rate)
        step_width_threshold = int(step_width_threshold * self.sampling_rate)

        # Gaussian smoothing of pressure signal
        smoothed_p = smoothing(signal, sigma)
        processed_pressure = smoothed_p

        # Get calibration points for each comment
        cal_times = []
        cal_volumes = []
        cal_pressures = []
        num_rejected = 0
        for idx in range(len(self.comment_idxs)):
            comment_idx = self.comment_idxs[idx]
            if make_plots:
                print('\n', self.time[comment_idx], self.comments[comment_idx])

            # check if there are duplicate comments for same calibration injection, only need one
            if (idx > 0) and (self.comment_idxs[idx] - self.comment_idxs[idx - 1] < dup_comment_distance):
                if make_plots:
                    print('Duplicate Comment for Same Calibration Injection.')
                continue

            # determine volume of calibration injection, which should be reported in ml
            vol_string = self.comments[comment_idx]
            new_string = ''
            for char in vol_string:
                if char in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '.']:
                    new_string += char
            try:
                cal_vol = float(new_string)
            except ValueError:
                if make_plots:
                    print('Unknown Injection Volume.')
                    num_rejected += 1
                continue

            if not (lowest_cal_vol <= cal_vol <= highest_cal_vol):
                # calibration volume not within expected range, discard this calibration injection
                if make_plots:
                    print('Injection Volume Outside Expected Range.')
                    num_rejected += 1
                continue

            # scale step window, w, based on injection volume (w = 200 / injection volume), then downsampling rate
            w = int((ref_cal_vol / cal_vol) * w_ref_cal_volume)
            w = int(w / downsampling_rate)

            # downsample data_idx array, making sure to include comment_idx
            ds_data_idx_left = self.data_idx[comment_idx::-downsampling_rate]
            ds_data_idx_left = ds_data_idx_left[::-1]  # reverse array
            ds_data_idx_right = self.data_idx[comment_idx::downsampling_rate]
            ds_data_idx = np.append(ds_data_idx_left[0:-1], ds_data_idx_right)  # only include comment_idx once

            # determine left and right indices of region of interest after downsampling
            left_idx = comment_idx - search_left
            right_idx = comment_idx + search_right
            ds_roi = np.array([idx for idx in ds_data_idx if left_idx <= idx <= right_idx])
            left_idx = ds_roi[0]
            right_idx = ds_roi[-1]

            # account for step window size, w
            possible_steps_roi = ds_roi[w:-w]
            if len(ds_roi) < (2 * w):
                # discard this calibration
                if make_plots:
                    print('Region of Interest Too Small.')
                    num_rejected += 1
                continue

            # Get ROI segments from different signals used for plots
            time_roi = self.data_idx[left_idx:right_idx]
            pressure_roi = self.pressure[left_idx:right_idx]
            # smoothed_p_roi = smoothed_p[left_idx:right_idx]
            processed_pressure_roi = processed_pressure[left_idx:right_idx]

            # Calculate variance within region of interest, don't downsample individual variance calculations
            # w_variance is window size for calculating signal variance, should be smaller than step window size, w
            signal_vars = []
            for var_idx in possible_steps_roi:
                variance = np.var(signal[var_idx - w_variance:var_idx + w_variance])
                signal_vars.append(variance)

            # Truncate region of interest if signal variance drops below lower threshold
            median_variance = np.median(signal_vars)
            lower_variance_threshold = median_variance * lower_variance_ratio
            start_idx = np.where(possible_steps_roi == comment_idx)[0][0]
            if signal_vars[start_idx] <= lower_variance_threshold:
                # discard this calibration
                if make_plots:
                    print('Comment Index Exceeds Lower Variance Threshold.')
                    num_rejected += 1
                continue
            vars_left_idx = 0
            vars_right_idx = len(possible_steps_roi) - 1

            search_idx = start_idx
            while search_idx <= vars_right_idx:
                if signal_vars[search_idx] <= lower_variance_threshold:
                    vars_right_idx = search_idx - 1
                search_idx += 1

            search_idx = start_idx
            while search_idx >= vars_left_idx:
                if signal_vars[search_idx] <= lower_variance_threshold:
                    vars_left_idx = search_idx + 1
                search_idx -= 1

            if (vars_right_idx - vars_left_idx) < (2 * w):
                # discard this calibration
                if make_plots:
                    print('Truncated Region of Interest Too Small.')
                    num_rejected += 1
                continue
            possible_steps_roi = possible_steps_roi[vars_left_idx:vars_right_idx + 1]

            # Moving Step Fit function (step presumed at middle of window, w, find linear fit parameters for left and
            # right sides of step (having same slope), f(x), as well as whole window g(x). Calculate residual sum of
            # squares for f(x) and g(x), take the difference to find steps (where [RSS(g) - RSS(f)]*h is large).
            # My adjustments: Assume the line fits will have slope of 0 because I want plateaus (can take average of
            # segments to get line parameter, b. Get residuals for segments and compare.
            step_probs = []
            for step_idx in possible_steps_roi:
                t_left = self.time[step_idx - w:step_idx:downsampling_rate]
                t_right = self.time[step_idx:step_idx + w:downsampling_rate]
                t_whole = self.time[step_idx - w:step_idx + w:downsampling_rate]
                left = processed_pressure[step_idx - w:step_idx:downsampling_rate]
                right = processed_pressure[step_idx:step_idx + w:downsampling_rate]
                whole = processed_pressure[step_idx - w:step_idx + w:downsampling_rate]

                # numpy.polyfit() works faster than hand computing mean, rss for more than 2000 data points
                left_coeff, left_rss = np.polyfit(t_left, left, 0, full=True)[0:2]
                right_coeff, right_rss = np.polyfit(t_right, right, 0, full=True)[0:2]
                whole_coeff, whole_rss = np.polyfit(t_whole, whole, 0, full=True)[0:2]
                left_mean = left_coeff[0]
                right_mean = right_coeff[0]
                step_h = right_mean - left_mean
                step_prob = (whole_rss - (left_rss + right_rss)) * step_h
                step_probs.append(step_prob)

            # Determine step threshold based on statistics of step_probs trace
            std_step_probs = np.std(step_probs)
            step_height_threshold = num_std_step_height_threshold * std_step_probs

            # Determine baseline and calibration plateau indices
            coords = self.moving_step_plateaus(comment_idx, possible_steps_roi, step_probs,
                                               step_height_threshold, step_width_threshold, make_plots)
            if make_plots:
                print("Coords:", coords)

            # Get change in pressure for each valid calibration
            if coords == tuple():
                num_rejected += 1
            else:
                base_mean = np.mean(signal[coords[0]:coords[1]])
                cal_mean = np.mean(signal[coords[2]:coords[3]])
                cal_pressure = cal_mean - base_mean
                cal_time = self.time[comment_idx]
                cal_times.append(cal_time)
                cal_volumes.append(cal_vol)
                cal_pressures.append(cal_pressure)

            if make_plots:
                # Plot step_probs for each calibration injection
                fig = plt.figure()
                ax1 = fig.add_subplot()
                ax1.set_xlabel('Time [s]')
                ax1.set_ylabel('Pressure [Pa]')
                ax1.tick_params(axis='y', labelcolor='k')
                ax2 = ax1.twinx()
                ax2.set_ylabel('Step Potential')
                ax2.tick_params(axis='y', labelcolor='r')
                ax3 = ax1.twinx()

                ax1.plot(time_roi, pressure_roi, 'k-')
                ax1.plot(time_roi, processed_pressure_roi, 'b-')
                ax1.annotate(self.comments[comment_idx], (self.data_idx[comment_idx], self.pressure[comment_idx]),
                             xycoords='data', xytext=(-30, 50), textcoords='offset pixels',
                             horizontalalignment='right', verticalalignment='top',
                             arrowprops=dict(arrowstyle='-|>', facecolor='r',
                                             connectionstyle='angle,angleA=0,angleB=90'))
                ax2.plot(possible_steps_roi, step_probs, 'r-')
                ax2.plot([time_roi[0], time_roi[-1]], [step_height_threshold, step_height_threshold], 'b--')
                ax2.plot([time_roi[0], time_roi[-1]], [-1 * step_height_threshold, -1 * step_height_threshold], 'b--')

                # ax3.plot(possible_steps_roi, signal_vars[vars_left_idx:vars_right_idx + 1], 'g-')

                if not coords == tuple():
                    ax2.plot([self.data_idx[coords[0]], self.data_idx[coords[0]]],
                             [step_height_threshold, -1 * step_height_threshold], 'm-')
                    ax2.plot([self.data_idx[coords[1]], self.data_idx[coords[1]]],
                             [step_height_threshold, -1 * step_height_threshold], 'm-')
                    ax2.plot([self.data_idx[coords[2]], self.data_idx[coords[2]]],
                             [step_height_threshold, -1 * step_height_threshold], 'm-')
                    ax2.plot([self.data_idx[coords[3]], self.data_idx[coords[3]]],
                             [step_height_threshold, -1 * step_height_threshold], 'm-')

                # fig.tight_layout()

        # Fit calibration points with line through origin (y = mx).
        # m = sum(x_i * y_i) / sum(x_i * x_i)
        numerator_sum = 0
        denominator_sum = 0
        if make_plots:
            print('\nCalibration Points:')
        for time, volume, pressure in zip(cal_times, cal_volumes, cal_pressures):
            if make_plots:
                print(time, volume, pressure)
            numerator_i = volume * pressure
            denominator_i = volume ** 2
            numerator_sum += numerator_i
            denominator_sum += denominator_i
        if denominator_sum == 0:
            slope = 0
        else:
            slope = numerator_sum / denominator_sum

        # Exclude outlier points using residuals for each point. If a point's residual is more than 1 standard
        # deviation away from the mean residual value, it is excluded.
        residuals = []
        for volume, pressure in zip(cal_volumes, cal_pressures):
            expected_pressure = slope * volume
            residual = (pressure - expected_pressure) ** 2
            residuals.append(residual)

        mean_residual = np.mean(residuals)
        std_residual = np.std(residuals)

        verified_cal_times = []
        verified_cal_volumes = []
        verified_cal_pressures = []
        excluded_cal_times = []
        excluded_cal_volumes = []
        excluded_cal_pressures = []
        for residual, time, volume, pressure in zip(residuals, cal_times, cal_volumes, cal_pressures):
            if residual <= mean_residual + num_std_exclude * std_residual:
                verified_cal_times.append(time)
                verified_cal_volumes.append(volume)
                verified_cal_pressures.append(pressure)
            else:
                excluded_cal_times.append(time)
                excluded_cal_volumes.append(volume)
                excluded_cal_pressures.append(pressure)

        # Recalculate slope of fit line with excluded points removed
        numerator_sum = 0
        denominator_sum = 0
        if make_plots:
            print('\nVerified Calibration Points:')
        for time, volume, pressure in zip(verified_cal_times, verified_cal_volumes, verified_cal_pressures):
            if make_plots:
                print(time, volume, pressure)
            numerator_i = volume * pressure
            denominator_i = volume ** 2
            numerator_sum += numerator_i
            denominator_sum += denominator_i
        if denominator_sum == 0:
            verified_slope = 0
        else:
            verified_slope = numerator_sum / denominator_sum

        # Calculate r_squared value
        mean_pressure = np.mean(verified_cal_pressures)
        unexplained_error = 0
        total_error = 0
        for volume, pressure in zip(verified_cal_volumes, verified_cal_pressures):
            expected_pressure = verified_slope * volume
            unexplained_error += (pressure - expected_pressure) ** 2
            total_error += (pressure - mean_pressure) ** 2
        r_squared = 1 - (unexplained_error / total_error)

        if make_plots:
            fig2 = plt.figure()
            ax = fig2.add_subplot()
            ax.set_xlabel('Change in Volume [ml]')
            ax.set_ylabel('Change in Pressure [Pa]')
            line1, = ax.plot(excluded_cal_volumes, excluded_cal_pressures, 'ro')
            ax.plot([0, highest_cal_vol], [0, slope * highest_cal_vol], 'r--')
            line2, = ax.plot(verified_cal_volumes, verified_cal_pressures, 'ko')
            line3, = ax.plot([0, highest_cal_vol], [0, verified_slope * highest_cal_vol], 'k-')
            line1.set_label('%s Excluded Points' % len(excluded_cal_volumes))
            line2.set_label('%s Included Points' % len(verified_cal_volumes))
            line3.set_label('y = %.6f x\nr^2 = %.6f' % (verified_slope, r_squared))
            ax.legend(loc='upper left')
            plt.show()

        # Add final calibration points and linear fit parameters to data arrays
        for time, volume, pressure in zip(verified_cal_times, verified_cal_volumes, verified_cal_pressures):
            self.add_calibration_point(time, pressure, volume)
        for time, volume, pressure in zip(excluded_cal_times, excluded_cal_volumes, excluded_cal_pressures):
            self.add_excluded_calibration_point(time, pressure, volume)
        self.excluded_calibration_m = slope
        self.calibration_m = verified_slope
        self.calibration_b = 0
        self.calibration_r_squared = r_squared

        print("\nCalibration Summary\nIncluded: %s, Excluded: %s, Rejected: %s"
              % (len(verified_cal_volumes), len(excluded_cal_volumes), num_rejected))

        return

    def moving_step_plateaus(self, comment_idx, possible_step_idxs, possible_step_probs,
                             step_height_threshold, step_width_threshold, make_plots=False):
        # Logic to select calibration and then baseline plateaus

        # First, check if comment_idx step_prob is within step_thresholds, if not, discard calibration
        start_idx = np.where(possible_step_idxs == comment_idx)[0][0]
        if abs(possible_step_probs[start_idx]) >= step_height_threshold:
            if make_plots:
                print('Comment Index Exceeds Step Height Threshold.')
            return tuple()

        # Look to right of comment index for end of calibration plateau
        search_idx = start_idx + 1
        cal_right_idx = len(possible_step_idxs) - 1
        while search_idx < len(possible_step_idxs):
            current_step_prob = possible_step_probs[search_idx]
            if abs(current_step_prob) >= step_height_threshold:
                # Step detected
                cal_right_idx = search_idx - 1
                break
            search_idx += 1

        # print("\nSearch for cal_right:")
        # print("step_found:", step_found)
        # print("cal_right_idx:", cal_right_idx)
        # print("min_step_prob:", min_step_prob)
        # print("max_step_prob:", max_step_prob)

        # Look to left of comment index for start of calibration plateau and end of baseline plateau
        search_idx = start_idx - 1
        cal_left_idx = 0
        base_right_idx = 0
        step_found = False
        while search_idx >= 0:
            current_step_prob = possible_step_probs[search_idx]
            if current_step_prob >= step_height_threshold:
                if step_found:
                    # continue searching for end of step
                    base_right_idx = search_idx
                else:
                    # Found start of upward step
                    step_found = True
                    cal_left_idx = search_idx + 1
                    base_right_idx = search_idx
            elif current_step_prob <= (-1) * step_height_threshold:
                # End search and discard calibration; step should move upward from baseline to calibration plateau
                if make_plots:
                    print('Search for cal_left_idx failed. Step from baseline to calibration in wrong direction.')
                return tuple()
            elif step_found:
                # End search if upward step has been found and within step_height_threshold again
                base_right_idx = search_idx
                break
            search_idx -= 1

        # print("\nSearch for cal_left/base_right:")
        # print("step_found:", step_found)
        # print("cal_left_idx:", cal_left_idx)
        # print("base_right_idx:", base_right_idx)
        # print("min_step_prob:", min_step_prob)
        # print("max_step_prob:", max_step_prob)

        # Continue looking left to search for start of baseline plateau
        search_idx = base_right_idx - 1
        base_left_idx = 0
        while search_idx >= 0:
            current_step_prob = possible_step_probs[search_idx]
            if abs(current_step_prob) >= step_height_threshold:
                # Step Detected
                base_left_idx = search_idx + 1
                break
            search_idx -= 1

        # print("\nSearch for base_left:")
        # print("step_found:", step_found)
        # print("base_left_idx:", base_left_idx)
        # print("min_step_prob:", min_step_prob)
        # print("max_step_prob:", max_step_prob)

        # Make sure each plateau is longer than step_width_threshold
        base_left_idx = possible_step_idxs[base_left_idx]
        base_right_idx = possible_step_idxs[base_right_idx]
        cal_left_idx = possible_step_idxs[cal_left_idx]
        cal_right_idx = possible_step_idxs[cal_right_idx]
        base_length = base_right_idx - base_left_idx
        cal_length = cal_right_idx - cal_left_idx

        # print("\nPlateau width check:")
        # print("base_length:", base_length)
        # print("cal_length:", cal_length)

        if base_length < step_width_threshold:
            if make_plots:
                print('Baseline plateau length (%ss) does not reach step width threshold'
                      % (base_length / self.sampling_rate))
            return tuple()
        if cal_length < step_width_threshold:
            if make_plots:
                print('Calibration plateau length (%ss) does not reach step width threshold'
                      % (cal_length / self.sampling_rate))
            return tuple()

        # Return indices defining left and right ends of baseline and calibration plateaus, in order
        return tuple([base_left_idx, base_right_idx, cal_left_idx, cal_right_idx])


def butter_coefficients(cutoff, fs, btype, order=5):
    nyq = 0.5 * fs
    normal_cutoff = cutoff / nyq
    sos = butter(order, normal_cutoff, btype=btype, analog=False, output='sos')
    return sos


def butter_filter(data, cutoff, fs, btype, order=5):
    sos = butter_coefficients(cutoff, fs, btype, order)
    filtered = sosfiltfilt(sos, data)
    return filtered


def find_zero_crossing(window, start_idx, find_inhale_onset, zero_cross_threshold):
    if find_inhale_onset:
        # Find inhale onset
        possible_onsets_from_left = [idx for idx, x in enumerate(window) if x >= zero_cross_threshold]
        possible_onsets_from_right = [len(window) - idx for idx, x in enumerate(np.flip(window))
                                      if x < zero_cross_threshold]

        if possible_onsets_from_left and possible_onsets_from_right:
            # At least one zero cross within window
            left_onset = possible_onsets_from_left[0]
            right_onset = possible_onsets_from_right[0]
            if left_onset == right_onset:
                # Only one zero crossing; inhale starts here
                onset_idx = left_onset
            else:
                # Multiple zero crossings.
                num_less = window[np.where(window[left_onset:right_onset] < zero_cross_threshold)].size
                num_greater = window[np.where(window[left_onset:right_onset] >= zero_cross_threshold)].size
                if num_less > num_greater:
                    # Count segment between outer zero crossings as part of exhale
                    onset_idx = right_onset
                elif num_greater > num_less:
                    # Count segment as part of inhale
                    onset_idx = left_onset
                else:
                    # For ties, count segment as part of inhale
                    onset_idx = left_onset

        elif possible_onsets_from_left:
            # No zero cross within window; entire window must be at or above zero_cross_threshold
            onset_idx = 0
        elif possible_onsets_from_right:
            # No zero cross within window; entire window must be below zero_cross_threshold
            onset_idx = len(window) - 1
        else:
            # This case should never be examined
            onset_idx = -1
            pass

    else:
        # Find exhale onset
        possible_onsets_from_left = [idx for idx, x in enumerate(window) if x < zero_cross_threshold]
        possible_onsets_from_right = [len(window) - idx for idx, x in enumerate(np.flip(window))
                                      if x >= zero_cross_threshold]

        if possible_onsets_from_left and possible_onsets_from_right:
            # At least one zero cross within window
            left_onset = possible_onsets_from_left[0]
            right_onset = possible_onsets_from_right[0]
            if left_onset == right_onset:
                # Only one zero crossing; exhale starts here
                onset_idx = left_onset
            else:
                # Multiple zero crossings.
                num_less = window[np.where(window[left_onset:right_onset] < zero_cross_threshold)].size
                num_greater = window[np.where(window[left_onset:right_onset] >= zero_cross_threshold)].size
                if num_less > num_greater:
                    # Count segment between outer zero crossings as part of exhale
                    onset_idx = left_onset
                elif num_greater > num_less:
                    # Count segment as part of inhale
                    onset_idx = right_onset
                else:
                    # For ties, count segment as part of exhale
                    onset_idx = left_onset

        elif possible_onsets_from_left:
            # No zero cross within window; entire window must be below zero_cross_threshold
            onset_idx = 0
        elif possible_onsets_from_right:
            # No zero cross within window; entire window must be at or above zero_cross_threshold
            onset_idx = len(window) - 1
        else:
            # This case should never be examined
            onset_idx = -1
            pass

    # Adjust indices to those of entire data set
    onset = start_idx + onset_idx
    pause_onset = -1

    return onset, pause_onset


def get_onsets_from_hist(flow_data, volume_data, start_idx, stop_idx, volume_threshold_idx, find_inhale_onset,
                         zero_cross_threshold, n_bins, extra_zero_bin_threshold, extra_pause_bin_threshold,
                         mode_bin_ratio, extra_pause_bin_ratio, find_pause=True, show_hist=False):

    # Determine upper and lower bin thresholds. Pause must occur within these values.
    max_extra_zero_bins = round(extra_zero_bin_threshold * n_bins)
    max_extra_pause_bins = round(extra_pause_bin_threshold * n_bins)

    window = flow_data[start_idx:stop_idx]
    vol_window = volume_data[start_idx:stop_idx]
    hist, bin_edges = np.histogram(window, n_bins)
    hist_mean_value = np.mean(hist)

    if show_hist:
        # Plot Histogram
        fig = plt.figure()
        ax = fig.add_subplot()
        ax.plot(bin_edges[1::], hist)
        fig.show()

    # Find bin where flow = 0
    zero_bin = int(np.floor(len(hist) / 2))  # initialize zero bin to center of histogram
    for idx in range(1, len(bin_edges)):
        if bin_edges[idx] > 0:
            zero_bin = idx - 1
            break

    # Find mode bin within range of zero_bin +/- max_extra_zero_bins
    mode_bin_left = zero_bin - max_extra_zero_bins
    mode_bin_right = zero_bin + max_extra_zero_bins
    if mode_bin_left < 0:
        mode_bin_left = 0
    if mode_bin_right > n_bins - 1:
        mode_bin_right = n_bins - 1

    rel_mode_bin = np.argmax(hist[mode_bin_left:mode_bin_right + 1])
    mode_bin = rel_mode_bin + mode_bin_left

    # Check if mode_bin value is greater than threshold
    mode_bin_value = hist[mode_bin]

    # Troubleshooting
    troubleshooting = False
    if find_pause and troubleshooting:
        print('start, stop idx:', start_idx, stop_idx)
        print('hist:', hist)
        print('bin_edges:', bin_edges)
        print('zero_bin:', zero_bin)
        print('mode_bin_left:', mode_bin_left)
        print('mode_bin_right:', mode_bin_right)
        print('mode_bin:', mode_bin)
        print('mode_bin_value:', mode_bin_value)

    if find_pause and (mode_bin_value >= mode_bin_ratio * hist_mean_value):
        # Pause detected, add adjacent bins if appropriate
        pause_bin_left = mode_bin
        pause_bin_right = mode_bin
        for distance in range(1, max_extra_pause_bins + 1):
            left_idx = mode_bin - distance
            right_idx = mode_bin + distance
            if (left_idx in range(len(hist))) and (hist[left_idx] >= extra_pause_bin_ratio * mode_bin_value):
                pause_bin_left = left_idx
            if (right_idx in range(len(hist))) and (hist[right_idx] >= extra_pause_bin_ratio * mode_bin_value):
                pause_bin_right = right_idx

        # Find onsets.
        pause_range_min_value = bin_edges[pause_bin_left]
        pause_range_max_value = bin_edges[pause_bin_right + 1]
        possible_pause = [idx + start_idx for idx, x in enumerate(window)
                          if pause_range_min_value <= x < pause_range_max_value]

        if possible_pause[0] <= volume_threshold_idx:
            # Pause starts before volume threshold exceeded. Count as pause onset
            first_pause_idx = possible_pause[0]

            # Find last pause idx before volume threshold. If threshold idx falls within pause segment, complete segment
            if volume_threshold_idx in possible_pause:
                # Look to right of volume_threshold_idx to find end of current pause segment
                search_idx = np.where(possible_pause == volume_threshold_idx)[0][0]
                last_pause_idx = possible_pause[search_idx]
                searching = True
                while searching and (search_idx < len(possible_pause) - 1):
                    last_pause_idx = possible_pause[search_idx]
                    search_idx = search_idx + 1
                    current_pause_idx = possible_pause[search_idx]
                    if current_pause_idx - last_pause_idx > 1:
                        searching = False

            else:
                # Look to left of volume_threshold_idx to find end of previous pause segment
                last_pause_idx = possible_pause[np.where(possible_pause < volume_threshold_idx)[0][-1]]

            onset = last_pause_idx + 1
            pause_onset = first_pause_idx

        else:
            # Pause starts after volume threshold exceeded. Use zero crossing to find onset
            onset, pause_onset = find_zero_crossing(window, start_idx, find_inhale_onset, zero_cross_threshold)

    else:
        # No pause, use zero crossing to find onset. Check from left and right; if these indices are not the same,
        # look between them. If most points have positive value, consider section part of inhale, otherwise exhale
        onset, pause_onset = find_zero_crossing(window, start_idx, find_inhale_onset, zero_cross_threshold)

    return onset, pause_onset


def linear_regression(x_data, y_data):
    slope, intercept, r_value, p_value, std_err = stats.linregress(x_data, y_data)
    return slope, intercept, r_value, p_value, std_err


def pressure_to_volume(pressure, ref_pressure, calibration_m, calibration_b):
    change_in_pressure = pressure - ref_pressure
    return (change_in_pressure - calibration_b) / calibration_m


def smoothing(signal, sigma=2.0):
    smoothed = gaussian_filter1d(signal, sigma)
    return smoothed


def voltage_to_pressure(voltage, supply_voltage, min_pascals, max_pascals):
    """This function takes a single voltage value and converts it to the corresponding pressure value, based
       on the transfer function and properties of the pressure transducer used for signal collection"""
    return min_pascals + 1.25 * (voltage / supply_voltage - 0.1) * (max_pascals - min_pascals)


def main():
    pass


if __name__ == "__main__":
    main()
